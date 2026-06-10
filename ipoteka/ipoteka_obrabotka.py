import os
import re
import sys
import queue
import logging
import traceback
import threading
from pathlib import Path
from datetime import datetime

import pandas as pd
import xlrd
import customtkinter as ctk
from tkinter import filedialog
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

APP_NAME = "Ипотека_обработка"

# Путь к логотипу (будет встроен в exe через PyInstaller)
LOGO_FILENAME = "logo-sia.png"

EXE_DIR = Path(sys.executable).resolve().parent

# =========================
# НАСТРОЙКИ ЛОГИРОВАНИЯ
# =========================
LOGS_BASE_DIR = EXE_DIR / "Логи"
TODAY = datetime.now().strftime("%Y-%m-%d")
TIME_NOW = datetime.now().strftime("%H-%M-%S")
LOGS_TODAY_DIR = LOGS_BASE_DIR / f"Логи_{TODAY}"
LOGS_TODAY_DIR.mkdir(parents=True, exist_ok=True)

LOG_PATH = LOGS_TODAY_DIR / f"{APP_NAME}_{TODAY}_{TIME_NOW}.log"

OUTPUT_PATH = EXE_DIR / "txt_done.xlsx"

REGISTRY_PATH_DEFAULT = r"\\172.29.101.6\финансовый отдел\1  Управление финансами 2024-2025\18 Сопровождение кредитного портфеля\!!! Реестр договоров.xlsx"
REGISTRY_SHEET = "Реестр"
REGISTRY_BASE_PATH = Path(r"\\172.29.101.6\финансовый отдел\1  Управление финансами 2024-2025\18 Сопровождение кредитного портфеля")

BAD_WORDS = ["погашен", "закрыт", "недейств"]
GOOD_WORD = "график"
GRAPH_EXTS = {".xls", ".xlsx", ".xlsm", ".xlsb"}

TXT_SEPARATOR = ";"
TXT_COLUMNS = [0, 5, 6, 7]
TXT_RENAME = ["дата", "договор", "фио плательщика", "сумма"]

COL_FILE = "файл графика"
COL_MONTH = "месяц платежа"
COL_YEAR = "год платежа"
COL_TOTAL = "общий ежемесячный платеж"
COL_PRINCIPAL = "платеж в счет погашения основного долга"
COL_INTEREST = "платеж процентов"
COL_FIO_CONTRACT = "фио по договору"

logger = logging.getLogger(APP_NAME)


def get_resource_path(relative_path):
    """Возвращает правильный путь к файлу-ресурсу (работает в разработке и в собранном exe)"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class QueueLogHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        try:
            msg = self.format(record)
            self.log_queue.put(msg)
        except Exception:
            pass


def setup_logger(log_queue=None):
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    if log_queue is not None:
        queue_handler = QueueLogHandler(log_queue)
        queue_handler.setFormatter(formatter)
        logger.addHandler(queue_handler)


def normalize_text_value(value):
    if pd.isna(value):
        return ""
    s = str(value).strip()
    s = re.sub(r"\.0$", "", s)
    return s


def normalize_header(s):
    return str(s).strip().lower()


def normalize_path(path_value):
    if pd.isna(path_value):
        return None

    p = str(path_value).strip().strip('"')
    if not p:
        return None

    if p.lower().startswith("hyperlink"):
        q1 = p.find('"')
        q2 = p.rfind('"')
        if q1 != -1 and q2 != -1 and q2 > q1:
            p = p[q1 + 1:q2]
        else:
            parts = p.split(maxsplit=1)
            if len(parts) == 2:
                p = parts[1].strip('"')

    if p.lower().startswith("file:///"):
        p = p[8:].replace("/", "\\")

    return p


def extract_transh_number(filename):
    """
    Извлекает номер транша из имени файла для числовой сортировки.
    Работает с форматами:
    - график _1,2 транш_Барсукова.xls (найдет 1)
    - график_10 транш.xls (найдет 10)
    - СИА_график 3 транш Жукова.xls (найдет 3)
    - график платежа с 01.11.2024_1 транш.xls (найдет 1)
    - график _ 2 транш_Дитер.xls (найдет 2)
    """
    basename = os.path.basename(str(filename))
    match = re.search(r'(\d+)\s*транш', basename.lower())
    if match:
        return int(match.group(1))
    return float('inf')


def read_source_txt(txt_path: Path) -> pd.DataFrame:
    logger.info(f"Чтение исходного txt: {txt_path}")

    df = pd.read_csv(txt_path, sep=TXT_SEPARATOR, encoding="cp1251", header=None, dtype=str)
    logger.info(f"Исходный txt: строк={len(df)}, колонок={len(df.columns)}")

    if len(df.columns) <= max(TXT_COLUMNS):
        raise ValueError(
            f"В txt недостаточно колонок. Ожидались индексы {TXT_COLUMNS}, а фактически колонок: {len(df.columns)}"
        )

    df.columns = [str(i) for i in range(1, len(df.columns) + 1)]
    df = df.iloc[:, TXT_COLUMNS].copy()
    df = df.dropna(subset=[df.columns[2]])
    df.columns = TXT_RENAME

    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()

    df = df.apply(lambda col: col.str.upper())
    return df.reset_index(drop=True)


def load_main_registry(registry_path: Path) -> pd.DataFrame:
    if not registry_path.exists():
        raise FileNotFoundError(f"Не найден основной реестр: {registry_path}")

    logger.info(f"Чтение основного реестра: {registry_path}, лист: {REGISTRY_SHEET}")
    df = pd.read_excel(
        registry_path,
        sheet_name=REGISTRY_SHEET,
        engine="openpyxl",
        dtype=str
    )

    return df


def match_with_registry(df_txt: pd.DataFrame, df_registry: pd.DataFrame, registry_base_path: Path) -> pd.DataFrame:
    txt = df_txt.copy()
    reestr = df_registry.copy()

    txt["договор_norm"] = txt["договор"].map(normalize_text_value)
    reestr["Номер договора_norm"] = reestr["Номер договора"].map(normalize_text_value)

    merged = txt.merge(
        reestr[["Номер договора_norm", "Папка", "Вид", "ФИО"]].copy(),
        left_on="договор_norm",
        right_on="Номер договора_norm",
        how="left"
    )

    def build_registry_link(x):
        folder_part = normalize_text_value(x)
        if not folder_part:
            return None
        return str(registry_base_path / folder_part)

    merged["ссылка"] = merged["Папка"].apply(build_registry_link)
    merged["вид"] = merged["Вид"]
    merged["фио по реестру"] = merged["ФИО"]

    merged = merged.drop(
        columns=["договор_norm", "Номер договора_norm", "Папка", "Вид", "ФИО"],
        errors="ignore"
    )

    return merged


def is_bad_name(name: str) -> bool:
    n = name.lower()
    return any(w in n for w in BAD_WORDS)


def is_good_graphic_name(name: str) -> bool:
    n = name.lower()
    return GOOD_WORD in n and not is_bad_name(n)


def find_graphic_files_in_dir(directory: Path):
    files = []

    if not directory.exists() or not directory.is_dir():
        return files

    for f in directory.rglob("*"):
        if f.is_file() and f.suffix.lower() in GRAPH_EXTS and is_good_graphic_name(f.name):
            files.append(str(f))

    # Сортировка по номеру транша (числовая)
    files.sort(key=extract_transh_number)

    return files


def process_graph_source(row: pd.Series):
    src = row.get("ссылка")
    contract = row.get("договор")

    # Разделитель для лога
    logger.info("=" * 80)
    logger.info(f">> НАЧАЛО ОБРАБОТКИ ДОГОВОРА {contract}")
    logger.info("=" * 80)

    if pd.isna(src) or not str(src).strip():
        logger.warning(f"Договор {contract}: ссылка пустая")
        logger.info("")  # Пустая строка для читаемости
        return []

    raw_src = str(src).strip()
    p = Path(raw_src)

    logger.info(f"Ссылка: {raw_src}")

    if p.is_file():
        logger.info(f"Тип: файл")
        new_row = row.copy()
        new_row[COL_FILE] = str(p)
        logger.info(f"Добавлен файл: {p.name}")
        logger.info("-" * 80)
        logger.info("")  # Пустая строка для читаемости
        return [new_row]

    if p.exists() and p.is_dir():
        logger.info(f"Тип: папка")
        files = find_graphic_files_in_dir(p)
        logger.info(f"Найдено файлов графиков: {len(files)}")

        if not files:
            logger.warning(f"Файлы графиков не найдены в папке")
            logger.info("-" * 80)
            logger.info("")  # Пустая строка для читаемости
            return []

        for f in files:
            transh_num = extract_transh_number(f)
            transh_display = transh_num if transh_num != float('inf') else "не найден"
            logger.info(f"  - {Path(f).name} (транш {transh_display})")

        out = []
        for f in files:
            new_row = row.copy()
            new_row[COL_FILE] = f
            out.append(new_row)

        logger.info("-" * 80)
        logger.info("")  # Пустая строка для читаемости
        return out

    logger.warning(f"Путь не существует или недоступен: {raw_src}")
    logger.info("-" * 80)
    logger.info("")  # Пустая строка для читаемости
    return []


def expand_graph_paths(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    not_found = 0

    logger.info("=" * 80)
    logger.info(">> ЭТАП: ПОИСК ФАЙЛОВ ГРАФИКОВ И РАЗВЕРТЫВАНИЕ СТРОК")
    logger.info("=" * 80)
    logger.info("")  # Пустая строка для читаемости

    for _, row in df.iterrows():
        expanded = process_graph_source(row)
        if expanded:
            # Сортируем expanded внутри договора по номеру транша
            expanded.sort(key=lambda x: extract_transh_number(x[COL_FILE]))
            rows.extend(expanded)
        else:
            row_copy = row.copy()
            row_copy[COL_FILE] = None
            rows.append(row_copy)
            not_found += 1

    out = pd.DataFrame(rows).reset_index(drop=True)
    logger.info("=" * 80)
    logger.info(f"После разворота путей строк: {len(out)}")
    logger.info(f"Строк без найденного графика: {not_found}")
    logger.info("=" * 80)
    logger.info("")  # Пустая строка для читаемости
    return out


def cell_is_actual_yellow(book, sheet, rowx, colx):
    if colx >= sheet.ncols:
        return False

    try:
        xf_idx = sheet.cell_xf_index(rowx, colx)
        xf = book.xf_list[xf_idx]
        bg = xf.background

        if bg.fill_pattern in (0, None):
            return False

        color_idx = bg.pattern_colour_index
        if color_idx in (None, 64):
            return False

        rgb = book.colour_map.get(color_idx)
        if not rgb or not isinstance(rgb, tuple) or len(rgb) != 3:
            return False

        r, g, b = rgb
        return r >= 200 and g >= 200 and b <= 120
    except Exception:
        return False


def row_is_yellow_xls(book, sheet, rowx):
    return all(cell_is_actual_yellow(book, sheet, rowx, colx) for colx in range(3, 8))


def extract_values_from_graph(path, verbose=True):
    book = xlrd.open_workbook(path, formatting_info=True)
    sheet = book.sheet_by_index(1)

    last_yellow_row = None

    for rx in range(sheet.nrows):
        if row_is_yellow_xls(book, sheet, rx):
            last_yellow_row = rx

    if last_yellow_row is None:
        if verbose:
            logger.warning(f"Жёлтые строки не найдены: {path}")
        return None, None, None, None, None

    data_row = last_yellow_row + 1

    if data_row >= sheet.nrows:
        if verbose:
            logger.warning(f"Следующая строка после жёлтой отсутствует: {path}")
        return None, None, None, None, None

    month = sheet.cell_value(data_row, 3) if sheet.ncols > 3 else None
    year = sheet.cell_value(data_row, 4) if sheet.ncols > 4 else None
    total = sheet.cell_value(data_row, 5) if sheet.ncols > 5 else None
    principal = sheet.cell_value(data_row, 6) if sheet.ncols > 6 else None
    interest = sheet.cell_value(data_row, 7) if sheet.ncols > 7 else None

    if verbose:
        logger.info(f"Извлечены данные:")
        logger.info(f"  - последняя закрашенная строка D:H = {last_yellow_row + 1}")
        logger.info(f"  - взята строка = {data_row + 1}")
        logger.info(f"  - месяц = {month}")
        logger.info(f"  - год = {year}")
        logger.info(f"  - общий платеж = {total}")
        logger.info(f"  - основной долг = {principal}")
        logger.info(f"  - проценты = {interest}")

    return month, year, total, principal, interest


def extract_fio_from_graph(path, verbose=True):
    book = xlrd.open_workbook(path, formatting_info=True)
    sheet = book.sheet_by_index(0)

    fio = sheet.cell_value(2, 2) if sheet.nrows > 2 and sheet.ncols > 2 else None

    if verbose:
        logger.info(f"ФИО по договору = {fio}")

    return fio


def enrich_with_graph_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    for col in [COL_MONTH, COL_YEAR, COL_TOTAL, COL_PRINCIPAL, COL_INTEREST, COL_FIO_CONTRACT]:
        if col not in df.columns:
            df[col] = None

    success_count = 0
    not_found_count = 0

    logger.info("=" * 80)
    logger.info(">> ЭТАП: ИЗВЛЕЧЕНИЕ ДАННЫХ ИЗ ФАЙЛОВ ГРАФИКОВ")
    logger.info("=" * 80)
    logger.info("")  # Пустая строка для читаемости

    for i, raw_path in df[COL_FILE].items():
        path = normalize_path(raw_path)

        if not path:
            logger.warning(f"Строка {i}: путь к файлу графика пустой")
            not_found_count += 1
            logger.info("")  # Пустая строка для читаемости
            continue

        logger.info("-" * 80)
        logger.info(f">> ОБРАБОТКА ФАЙЛА: {Path(path).name}")
        logger.info(f"Путь: {path}")

        if not os.path.exists(path):
            logger.warning(f"Файл не найден")
            not_found_count += 1
            logger.info("")  # Пустая строка для читаемости
            continue

        try:
            month, year, total, principal, interest = extract_values_from_graph(path, verbose=True)
            fio = extract_fio_from_graph(path, verbose=True)

            df.at[i, COL_MONTH] = month
            df.at[i, COL_YEAR] = year
            df.at[i, COL_TOTAL] = total
            df.at[i, COL_PRINCIPAL] = principal
            df.at[i, COL_INTEREST] = interest
            df.at[i, COL_FIO_CONTRACT] = fio

            success_count += 1
            logger.info(f">> Файл обработан успешно")
            logger.info("")  # Пустая строка для читаемости

        except Exception as e:
            logger.exception(f"ОШИБКА обработки: {e}")
            logger.info("")  # Пустая строка для читаемости

    logger.info("=" * 80)
    logger.info(f"Успешно извлечены значения из графиков: {success_count}")
    logger.info(f"Не обработано графиков: {not_found_count}")
    logger.info("=" * 80)
    logger.info("")  # Пустая строка для читаемости

    return df


def save_txt_done(df: pd.DataFrame, output_path: Path):
    df.to_excel(output_path, index=False, engine="openpyxl")


def normalize_month_value(value):
    if value is None:
        return None
    if pd.isna(value):
        return None

    s = str(value).strip()
    if not s:
        return None

    s = re.sub(r"\.0$", "", s)
    return s


def apply_month_mode_highlight(ws, month_col_idx, fill):
    values = []

    for row in range(2, ws.max_row + 1):
        v = normalize_month_value(ws.cell(row=row, column=month_col_idx).value)
        if v is not None:
            values.append(v)

    if not values:
        logger.info("Столбец 'месяц платежа' пустой, подсветка по моде не выполнялась")
        return

    modes = pd.Series(values, dtype="object").mode()

    if len(modes) != 1:
        logger.info("В столбце 'месяц платежа' несколько мод, подсветка по моде не выполнялась")
        return

    mode_value = normalize_month_value(modes.iloc[0])
    highlighted = 0

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=month_col_idx)
        current = normalize_month_value(cell.value)
        if current is None:
            continue
        if current != mode_value:
            cell.fill = fill
            highlighted += 1

    logger.info(f"Подсвечено ячеек в столбце 'месяц платежа', отличающихся от моды: {highlighted}")


def format_final_excel(file_path: Path):
    wb = load_workbook(file_path)
    ws = wb[wb.sheetnames[0]]

    new_order = [
        "дата",
        "договор",
        "вид",
        "фио плательщика",
        "фио по договору",
        "фио по реестру",
        "сумма",
        "месяц платежа",
        "год платежа",
        "общий ежемесячный платеж",
        "платеж в счет погашения основного долга",
        "платеж процентов",
        "файл графика",
        "ссылка"
    ]

    highlight_headers = {
        "месяц платежа",
        "год платежа",
        "общий ежемесячный платеж",
        "платеж в счет погашения основного долга",
        "платеж процентов"
    }

    grey_headers = {
        "дата",
        "договор",
        "вид",
        "фио плательщика",
        "фио по договору",
        "фио по реестру",
        "сумма",
        "файл графика",
        "ссылка"
    }

    link_headers = {"ссылка", "файл графика"}

    hdr_fill = PatternFill(fill_type="solid", fgColor="FCD5B4")
    grey_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value is not None else ""
        headers.append(val)

    header_map = {normalize_header(v): i + 1 for i, v in enumerate(headers)}

    for col_name in new_order:
        if col_name not in header_map:
            raise ValueError(f'Столбец "{col_name}" не найден')

    data = []
    data.append([headers[header_map[col] - 1] for col in new_order])

    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in new_order:
            old_col_idx = header_map[col]
            row_data.append(ws.cell(row=row, column=old_col_idx).value)
        data.append(row_data)

    ws.delete_cols(1, ws.max_column)

    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    new_header_map = {}
    for cell in ws[1]:
        val = str(cell.value).strip().lower() if cell.value is not None else ""
        new_header_map[val] = cell.column

    for name in highlight_headers:
        col_idx = new_header_map[name]
        ws.cell(row=1, column=col_idx).fill = hdr_fill

    for name in link_headers:
        col_idx = new_header_map[name]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                link = str(cell.value).strip()
                cell.hyperlink = link
                cell.value = os.path.basename(link)
                cell.style = "Hyperlink"

    for name in grey_headers:
        col_idx = new_header_map[name]
        ws.cell(row=1, column=col_idx).fill = grey_fill

    month_col_idx = new_header_map[COL_MONTH]
    apply_month_mode_highlight(ws, month_col_idx, hdr_fill)

    widths = {
        "A": 10,
        "B": 13,
        "C": 8,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 10,
        "H": 17,
        "I": 13,
        "J": 25,
        "K": 25,
        "L": 19,
        "M": 30,
        "N": 20
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(file_path)


def run_processing(txt_path: Path, registry_path: Path, output_path: Path):
    logger.info("=" * 80)
    logger.info(f">> СТАРТ {APP_NAME}")
    logger.info("=" * 80)
    logger.info(f"Выбранный txt: {txt_path}")
    logger.info(f"Итоговый файл: {output_path}")
    logger.info(f"Основной реестр: {registry_path}")
    logger.info(f"Лист реестра: {REGISTRY_SHEET}")
    logger.info("=" * 80)
    logger.info("")  # Одна пустая строка перед началом основных операций

    df_txt = read_source_txt(txt_path)
    logger.info(f"Прочитано строк из txt: {len(df_txt)}")

    df_registry = load_main_registry(registry_path)
    logger.info(f"Прочитано строк из реестра: {len(df_registry)}")

    df_merged = match_with_registry(df_txt, df_registry, REGISTRY_BASE_PATH)
    logger.info(f"После сопоставления с реестром строк: {len(df_merged)}")
    logger.info("")  # Пустая строка перед следующим этапом

    df_expanded = expand_graph_paths(df_merged)
    logger.info(f"После поиска файлов графиков строк: {len(df_expanded)}")
    logger.info("")  # Пустая строка перед следующим этапом

    save_txt_done(df_expanded, output_path)
    logger.info("Черновой txt_done.xlsx сохранён")
    logger.info("")  # Пустая строка перед следующим этапом

    df_for_enrich = pd.read_excel(output_path, dtype=object)
    df_enriched = enrich_with_graph_data(df_for_enrich)

    save_txt_done(df_enriched, output_path)
    logger.info("txt_done.xlsx дополнен данными из графиков")
    logger.info("")  # Пустая строка перед финальным форматированием

    format_final_excel(output_path)
    logger.info("Финальное форматирование txt_done.xlsx завершено")
    logger.info("")  # Пустая строка перед финальным итогом
    logger.info("=" * 80)
    logger.info(">> ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО")
    logger.info("=" * 80)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.log_queue = queue.Queue()
        setup_logger(self.log_queue)

        self.title(APP_NAME)
        self.geometry("980x760")
        self.minsize(900, 680)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.worker_thread = None

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(5, weight=1)

        self.build_ui()
        self.after(100, self.poll_log_queue)

    def build_ui(self):
        # Верхняя панель с заголовком и логотипом
        self.top_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.top_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        self.top_frame.grid_columnconfigure(0, weight=1)

        self.header = ctk.CTkLabel(
            self.top_frame,
            text=APP_NAME,
            font=ctk.CTkFont(size=24, weight="bold")
        )
        self.header.grid(row=0, column=0, sticky="w")

        # Логотип (сохраняем пропорции, высота 80px)
        logo_path = get_resource_path(LOGO_FILENAME)
        try:
            if os.path.exists(logo_path):
                logo_image = Image.open(logo_path)
                # Сохраняем пропорции, высота 80 пикселей
                new_height = 80
                original_width, original_height = logo_image.size
                new_width = int(original_width * (new_height / original_height))
                logo_ctk = ctk.CTkImage(light_image=logo_image, dark_image=logo_image, size=(new_width, new_height))
                self.logo_label = ctk.CTkLabel(self.top_frame, image=logo_ctk, text="")
                self.logo_label.grid(row=0, column=1, padx=(15, 0), sticky="e")
            else:
                logger.warning(f"Файл логотипа не найден: {logo_path}")
        except Exception as e:
            logger.warning(f"Не удалось загрузить логотип: {e}")

        self.frame_top = ctk.CTkFrame(self)
        self.frame_top.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        self.frame_top.grid_columnconfigure(1, weight=1)

        self.lbl_txt = ctk.CTkLabel(self.frame_top, text="Исходный TXT-файл")
        self.lbl_txt.grid(row=0, column=0, padx=12, pady=(12, 6), sticky="w")

        self.entry_txt = ctk.CTkEntry(self.frame_top)
        self.entry_txt.grid(row=0, column=1, padx=12, pady=(12, 6), sticky="ew")

        self.btn_txt = ctk.CTkButton(self.frame_top, text="Выбрать TXT", width=140, command=self.choose_txt)
        self.btn_txt.grid(row=0, column=2, padx=12, pady=(12, 6))

        self.lbl_registry = ctk.CTkLabel(self.frame_top, text="Основной реестр")
        self.lbl_registry.grid(row=1, column=0, padx=12, pady=6, sticky="w")

        self.entry_registry = ctk.CTkEntry(self.frame_top)
        self.entry_registry.grid(row=1, column=1, padx=12, pady=6, sticky="ew")
        self.entry_registry.insert(0, REGISTRY_PATH_DEFAULT)

        self.btn_registry = ctk.CTkButton(self.frame_top, text="Выбрать Excel", width=140, command=self.choose_registry)
        self.btn_registry.grid(row=1, column=2, padx=12, pady=6)

        self.lbl_output = ctk.CTkLabel(self.frame_top, text="Итоговый файл")
        self.lbl_output.grid(row=2, column=0, padx=12, pady=(6, 12), sticky="w")

        self.entry_output = ctk.CTkEntry(self.frame_top)
        self.entry_output.grid(row=2, column=1, padx=12, pady=(6, 12), sticky="ew")
        self.entry_output.insert(0, str(OUTPUT_PATH))

        self.btn_open_folder = ctk.CTkButton(self.frame_top, text="Открыть папку EXE", width=140, command=self.open_exe_folder)
        self.btn_open_folder.grid(row=2, column=2, padx=12, pady=(6, 12))

        self.frame_actions = ctk.CTkFrame(self)
        self.frame_actions.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        self.frame_actions.grid_columnconfigure(0, weight=0)
        self.frame_actions.grid_columnconfigure(1, weight=0)
        self.frame_actions.grid_columnconfigure(2, weight=1)

        self.btn_run = ctk.CTkButton(self.frame_actions, text="Запустить обработку", command=self.start_processing, width=180)
        self.btn_run.grid(row=0, column=0, padx=12, pady=12, sticky="w")

        self.btn_clear = ctk.CTkButton(self.frame_actions, text="Очистить лог", command=self.clear_log, width=140)
        self.btn_clear.grid(row=0, column=1, padx=12, pady=12, sticky="w")

        self.status_label = ctk.CTkLabel(self.frame_actions, text="Статус: готово", anchor="w")
        self.status_label.grid(row=0, column=2, padx=12, pady=12, sticky="ew")

        self.progress = ctk.CTkProgressBar(self)
        self.progress.grid(row=3, column=0, padx=20, pady=(0, 10), sticky="ew")
        self.progress.set(0)

        self.info_label = ctk.CTkLabel(
            self,
            text=f"txt_done.xlsx будет сохранён рядом с exe:\n{OUTPUT_PATH}",
            justify="left"
        )
        self.info_label.grid(row=4, column=0, padx=20, pady=(0, 10), sticky="w")

        self.log_box = ctk.CTkTextbox(self, wrap="word")
        self.log_box.grid(row=5, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.log_box.configure(state="disabled")

        # Настройка цветовых тегов для логов
        self.log_box.tag_config("info", foreground="white")
        self.log_box.tag_config("warning", foreground="yellow")
        self.log_box.tag_config("error", foreground="orange")
        self.log_box.tag_config("critical", foreground="red")

    def choose_txt(self):
        path = filedialog.askopenfilename(
            title="Выберите исходный TXT-файл",
            filetypes=[("TXT files", "*.txt"), ("All files", "*.*")]
        )
        if path:
            self.entry_txt.delete(0, "end")
            self.entry_txt.insert(0, path)

    def choose_registry(self):
        path = filedialog.askopenfilename(
            title="Выберите основной реестр",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xls"), ("All files", "*.*")]
        )
        if path:
            self.entry_registry.delete(0, "end")
            self.entry_registry.insert(0, path)

    def open_exe_folder(self):
        os.startfile(str(EXE_DIR))

    def clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def append_log(self, text):
        self.log_box.configure(state="normal")

        # Определяем цвет тега по содержимому (убрано "!!!" из условия)
        if "| ERROR |" in text or "ОШИБКА" in text.upper():
            tag = "error"
        elif "| WARNING |" in text or "WARNING" in text.upper():
            tag = "warning"
        elif "| CRITICAL |" in text or "КРИТИЧЕСКАЯ" in text.upper():
            tag = "critical"
        else:
            tag = "info"

        self.log_box.insert("end", text + "\n", tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.append_log(msg)

                if "СТАРТ" in msg:
                    self.progress.set(0.05)
                    self.status_label.configure(text="Статус: обработка")
                elif "Прочитано строк из txt" in msg:
                    self.progress.set(0.15)
                elif "Прочитано строк из реестра" in msg:
                    self.progress.set(0.30)
                elif "После сопоставления с реестром строк" in msg:
                    self.progress.set(0.45)
                elif "После поиска файлов графиков строк" in msg:
                    self.progress.set(0.60)
                elif "txt_done.xlsx дополнен данными из графиков" in msg:
                    self.progress.set(0.85)
                elif "Финальное форматирование" in msg:
                    self.progress.set(0.95)
                elif "ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО" in msg:
                    self.progress.set(1.0)
                    self.status_label.configure(text="Статус: готово")
                elif "ОШИБКА" in msg.upper() and "Критическая" in msg:
                    self.status_label.configure(text="Статус: ошибка")
        except queue.Empty:
            pass

        self.after(100, self.poll_log_queue)

    def set_controls_state(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        self.btn_run.configure(state=state)
        self.btn_clear.configure(state=state)
        self.btn_txt.configure(state=state)
        self.btn_registry.configure(state=state)
        self.entry_txt.configure(state=state)
        self.entry_registry.configure(state=state)

    def start_processing(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return

        txt_value = self.entry_txt.get().strip()
        registry_value = self.entry_registry.get().strip()
        output_value = self.entry_output.get().strip()

        if not txt_value:
            self.append_log("Ошибка: не выбран TXT-файл.")
            self.status_label.configure(text="Статус: ошибка")
            return

        txt_path = Path(txt_value)
        registry_path = Path(registry_value)
        output_path = Path(output_value)

        self.progress.set(0)
        self.status_label.configure(text="Статус: обработка")
        self.set_controls_state(False)

        self.worker_thread = threading.Thread(
            target=self.worker_run,
            args=(txt_path, registry_path, output_path),
            daemon=True
        )
        self.worker_thread.start()

    def worker_run(self, txt_path: Path, registry_path: Path, output_path: Path):
        try:
            run_processing(txt_path, registry_path, output_path)
        except Exception as e:
            logger.error("=" * 80)
            logger.error(">> КРИТИЧЕСКАЯ ОШИБКА ВЫПОЛНЕНИЯ")
            logger.error("=" * 80)
            logger.error(str(e))
            logger.error(traceback.format_exc())
            self.progress.set(0)
            self.status_label.configure(text="Статус: ошибка")
        finally:
            self.after(0, lambda: self.set_controls_state(True))


if __name__ == "__main__":
    app = App()
    app.mainloop()