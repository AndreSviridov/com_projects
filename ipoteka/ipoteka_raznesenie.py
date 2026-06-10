import os
import sys
import queue
import shutil
import logging
import traceback
import threading
import re
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime
from tkinter import messagebox, filedialog, simpledialog

import pandas as pd
import customtkinter as ctk
from PIL import Image
from openpyxl import load_workbook
from openpyxl.comments import Comment

# =========================
# НАСТРОЙКИ ПРИЛОЖЕНИЯ
# =========================
APP_NAME = "Ипотека_разнесение"
LOGO_FILENAME = "logo-sia.png"

# Определяем пути для работы в режиме разработки и в собранном exe
def get_resource_path(relative_path):
    """Возвращает правильный путь к файлу-ресурсу (работает в разработке и в собранном exe)"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


EXE_DIR = Path(sys.executable).resolve().parent if getattr(sys, 'frozen', False) else Path.cwd()

# Создаём структуру папок для логов
LOGS_BASE_DIR = EXE_DIR / "Логи"
TODAY = datetime.now().strftime("%Y-%m-%d")
TIME_NOW = datetime.now().strftime("%H-%M-%S")
LOGS_TODAY_DIR = LOGS_BASE_DIR / f"Логи_{TODAY}"
LOGS_TODAY_DIR.mkdir(parents=True, exist_ok=True)

LOG_PATH = LOGS_TODAY_DIR / f"{APP_NAME}_{TODAY}_{TIME_NOW}.log"

# Папка для резервных копий
BACKUP_BASE_DIR = EXE_DIR / "резервные_копии"
BACKUP_BASE_DIR.mkdir(parents=True, exist_ok=True)

# Пути по умолчанию
BASE_PATH = r"\\172.29.101.6\финансовый отдел\1  Управление финансами 2024-2025\18 Сопровождение кредитного портфеля"

DEFAULT_PATHS = {
    "ИЖС": Path(BASE_PATH) / "оплата ИЖС_2025.xlsx",
    "ДДУ": Path(BASE_PATH) / "оплата ДДУ_2025.xlsx",
    "ЖСК": Path(BASE_PATH) / "оплата ЖСК_2025.xlsx",
    "основной": Path(BASE_PATH) / "!!! Реестр договоров.xlsx"
}

STATUS_COL_NAME = "Статус"
STATUS_DEBTOR = "должник"

MONTH_NAMES_RU = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
}

# =========================
# ЛОГИРОВАНИЕ
# =========================
logger = logging.getLogger(APP_NAME)


class QueueLogHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        try:
            msg = self.format(record)
            self.log_queue.put(msg)
        except Exception:
            pass


def setup_logger(log_queue=None):
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    if log_queue is not None:
        queue_handler = QueueLogHandler(log_queue)
        queue_handler.setFormatter(formatter)
        logger.addHandler(queue_handler)


# =========================
# ФУНКЦИИ РЕЗЕРВНОГО КОПИРОВАНИЯ
# =========================
def create_backup(source_files: dict) -> Path | None:
    """
    Создаёт резервную копию указанных файлов.
    source_files: dict с именами и путями к файлам
    Возвращает путь к папке с резервной копией или None при ошибке
    """
    backup_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_dir = BACKUP_BASE_DIR / backup_time
    backup_dir.mkdir(parents=True, exist_ok=True)

    logger.info("=" * 100)
    logger.info(">>> СОЗДАНИЕ РЕЗЕРВНЫХ КОПИЙ")
    logger.info("=" * 100)

    for name, file_path in source_files.items():
        if file_path and file_path.exists():
            backup_file = backup_dir / f"{name}_backup.xlsx"
            try:
                shutil.copy2(file_path, backup_file)
                logger.info(f"  Копия создана: {name} -> {backup_file.name}")
            except Exception as e:
                logger.error(f"  Ошибка копирования {name}: {e}")
                return None
        else:
            logger.warning(f"  Файл {name} не найден, пропущен")

    logger.info(f"Резервные копии сохранены в: {backup_dir}")
    logger.info("=" * 100)
    return backup_dir


def get_last_backup() -> Path | None:
    """Возвращает путь к последней папке с резервной копией или None"""
    backups = sorted(BACKUP_BASE_DIR.iterdir())
    backups = [b for b in backups if b.is_dir()]
    if not backups:
        return None
    backups.sort(key=lambda x: x.stat().st_ctime, reverse=True)
    return backups[0]


def get_all_backups() -> list[Path]:
    """Возвращает список всех папок с резервными копиями, отсортированный по убыванию даты"""
    backups = [b for b in BACKUP_BASE_DIR.iterdir() if b.is_dir()]
    backups.sort(key=lambda x: x.stat().st_ctime, reverse=True)
    return backups


def restore_from_backup(backup_dir: Path, target_files: dict) -> bool:
    """
    Восстанавливает файлы из резервной копии.
    backup_dir: папка с резервной копией
    target_files: словарь с целевыми путями {имя: путь}
    """
    logger.info("=" * 100)
    logger.info(">>> ВОССТАНОВЛЕНИЕ ИЗ РЕЗЕРВНОЙ КОПИИ")
    logger.info("=" * 100)
    logger.info(f"Резервная копия: {backup_dir}")

    success = True
    for name, target_path in target_files.items():
        backup_file = backup_dir / f"{name}_backup.xlsx"
        if backup_file.exists():
            try:
                shutil.copy2(backup_file, target_path)
                logger.info(f"  Восстановлен: {name} -> {target_path}")
            except Exception as e:
                logger.error(f"  Ошибка восстановления {name}: {e}")
                success = False
        else:
            logger.warning(f"  Файл {name}_backup.xlsx не найден в резервной копии")

    if success:
        logger.info("Восстановление завершено успешно!")
    else:
        logger.error("Восстановление завершено с ошибками")

    logger.info("=" * 100)
    return success


# =========================
# НОВЫЕ ФУНКЦИИ ДЛЯ СУММИРОВАНИЯ С ФОРМУЛОЙ И НАКОПЛЕНИЯ ДАТ
# =========================
def extract_formula_from_comment(cell):
    """
    Извлекает строку формулы из комментария ячейки.
    Возвращает None, если комментария нет.
    """
    if cell.comment and cell.comment.text:
        return cell.comment.text.strip()
    return None


def add_to_cell_with_formula(cell, amount, logger_prefix: str = ""):
    """
    Суммирует значение в ячейке с сохранением полной истории в комментарии.
    
    Логика:
    - Если ячейка пустая: записываем число
    - Если в ячейке реальный ноль: записываем число (ноль не считается историей)
    - Если есть комментарий: читаем историю, добавляем новое слагаемое
    - Если числа без комментария: превращаем в формулу =старое+новое
    """
    new_amount = to_decimal(amount)
    
    # Проверяем, есть ли комментарий с историей
    history = extract_formula_from_comment(cell)
    
    if history:
        # Уже есть формула с историей
        new_history = f"{history}+{new_amount}"
        formula = f"={new_history}"
        cell.value = formula
        cell.comment = Comment(new_history, "Ипотека_разнесение")
        logger.info(f"{logger_prefix} Обновлена формула: {formula}")
        return to_decimal(new_amount)
    
    # Нет истории, проверяем текущее значение
    # Проверяем, пустая ли ячейка
    if cell.value is None or str(cell.value).strip() == "":
        # Ячейка пустая — записываем просто число
        cell.value = float(new_amount)
        logger.info(f"{logger_prefix} Установлено число: {new_amount}")
        return new_amount
    
    # Проверяем, является ли текущее значение нулём (числовым нулём)
    current = to_decimal(cell.value)
    if current == 0:
        # В ячейке реальный ноль (число) — просто записываем новое число
        cell.value = float(new_amount)
        logger.info(f"{logger_prefix} Установлено число (был ноль): {new_amount}")
        return new_amount
    
    # Есть число (не ноль) без истории — превращаем в формулу
    # Округляем до 2 десятичных знаков для красоты
    current_rounded = round(current, 2)
    new_amount_rounded = round(new_amount, 2)
    new_history = f"{current_rounded}+{new_amount_rounded}"
    formula = f"={new_history}"
    cell.value = formula
    cell.comment = Comment(new_history, "Ипотека_разнесение")
    logger.info(f"{logger_prefix} Создана формула: {formula} (было число {current_rounded})")
    return current + new_amount


def format_date_for_display(date_value) -> str:
    """
    Приводит дату к формату ДД-ММ-ГГГГ.
    Поддерживает форматы:
    - datetime объект
    - YYYY-MM-DD HH:MM:SS
    - DD-MM-YYYY
    - DD.MM.YYYY
    """
    if date_value is None or str(date_value).strip() == "":
        return None
    
    # Если это datetime объект
    if isinstance(date_value, datetime):
        return date_value.strftime("%d-%m-%Y")
    
    # Если это строка
    date_str = str(date_value).strip()
    
    # Формат YYYY-MM-DD HH:MM:SS
    match = re.match(r'(\d{4})-(\d{2})-(\d{2})', date_str)
    if match:
        year, month, day = match.groups()
        return f"{day}-{month}-{year}"
    
    # Формат DD-MM-YYYY (уже правильный)
    match = re.match(r'(\d{2})-(\d{2})-(\d{4})', date_str)
    if match:
        return date_str
    
    # Формат DD.MM.YYYY
    match = re.match(r'(\d{2})\.(\d{2})\.(\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        return f"{day}-{month}-{year}"
    
    # Если ничего не подошло, возвращаем как есть
    return date_str


def append_date_to_cell(cell, new_date, logger_prefix: str = ""):
    """
    Добавляет новую дату к существующим через запятую.
    Все даты приводятся к формату ДД-ММ-ГГГГ.
    Новая дата добавляется в конец списка (хронологический порядок).
    """
    formatted_new_date = format_date_for_display(new_date)
    if formatted_new_date is None:
        return
    
    current = cell.value
    
    if current is None or str(current).strip() == "":
        # Ячейка пустая
        cell.value = formatted_new_date
        logger.info(f"{logger_prefix} Установлена дата: {formatted_new_date}")
        return
    
    # Форматируем текущее значение
    current_str = str(current).strip()
    
    # Если в текущем значении несколько дат через запятую
    if ',' in current_str:
        dates = [d.strip() for d in current_str.split(',')]
        formatted_dates = [format_date_for_display(d) or d for d in dates]
    else:
        formatted_dates = [format_date_for_display(current_str) or current_str]
    
    # Добавляем новую дату, если её ещё нет
    if formatted_new_date not in formatted_dates:
        formatted_dates.append(formatted_new_date)
    
    # Формируем результат
    new_value = ', '.join(formatted_dates)
    cell.value = new_value
    logger.info(f"{logger_prefix} Добавлена дата: {new_value}")


# =========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================
def normalize_text(value):
    if value is None:
        return ""
    return str(value).replace(" ", "").strip()


def cell_contains_minus(value) -> bool:
    if value is None:
        return False
    return '-' in str(value)


def to_decimal(value) -> Decimal:
    if value is None:
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if text == "":
        return Decimal('0')
    text = text.replace(" ", "").replace("\xa0", "")
    if ',' in text and '.' in text:
        text = text.replace('.', '').replace(',', '.')
    elif ',' in text:
        text = text.replace(',', '.')
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal('0')


def ensure_status_column_in_df(df: pd.DataFrame) -> pd.DataFrame:
    if STATUS_COL_NAME not in df.columns:
        df[STATUS_COL_NAME] = ""
    return df


def load_txt_done_df(file_path: str) -> pd.DataFrame:
    df = pd.read_excel(file_path)
    df = ensure_status_column_in_df(df)
    return df


def mark_rows_as_debtor(df: pd.DataFrame, mask, logger_prefix: str = ""):
    count = int(mask.sum())
    if count > 0:
        df.loc[mask, STATUS_COL_NAME] = STATUS_DEBTOR
        logger.warning(f"{logger_prefix} !!! В txt_done.xlsx проставлен статус '{STATUS_DEBTOR}' для строк: {count}")


def is_debtor_status(value) -> bool:
    return str(value).strip().lower() == STATUS_DEBTOR


def mark_debtor_in_excel_directly(file_path: str, df: pd.DataFrame, mask, status_value: str, logger_prefix: str = ""):
    excel_rows = [idx + 2 for idx, val in enumerate(mask) if val]
    if not excel_rows:
        logger.info(f"{logger_prefix} Нет строк для простановки статуса '{status_value}'")
        return

    logger.info(f"{logger_prefix} Простановка статуса '{status_value}' в Excel: строки {excel_rows}")

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        status_col = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).strip() == STATUS_COL_NAME:
                status_col = col
                break

        if status_col is None:
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col, value=STATUS_COL_NAME)
            logger.info(f"{logger_prefix} Столбец '{STATUS_COL_NAME}' не найден, добавлен в колонку {status_col}")

        for row in excel_rows:
            ws.cell(row=row, column=status_col, value=status_value)

        wb.save(file_path)
        wb.close()
        logger.info(f"{logger_prefix} Статус '{status_value}' успешно записан в {len(excel_rows)} строк")

    except Exception as e:
        logger.error(f"{logger_prefix} Ошибка при записи статуса в Excel: {e}")
        raise


# =========================
# ФУНКЦИИ ДЛЯ ИЖС/ДДУ
# =========================
def get_tranche_from_filename(filename: str) -> str | None:
    fname = str(Path(filename).name).lower()
    if ('график _ 1' in fname or 'график _1' in fname or 'график 1' in fname) and 'график_10' not in fname and 'график_11' not in fname and 'график_1,2' not in fname:
        return '1'
    if ('график _ 2' in fname or 'график _2' in fname or 'график 2' in fname) and 'график_1,2' not in fname and 'график_2,3' not in fname:
        return '2'
    if ('график _ 3' in fname or 'график _3' in fname or 'график 3' in fname) and 'график_2,3' not in fname:
        return '3'
    if 'график _ 4' in fname or 'график _4' in fname or 'график 4' in fname:
        return '4'
    if 'график _ 5' in fname or 'график _5' in fname or 'график 5' in fname:
        return '5'
    if 'график _ 6' in fname or 'график _6' in fname or 'график 6' in fname:
        return '6'
    if 'график _ 7' in fname or 'график _7' in fname or 'график 7' in fname:
        return '7'
    if 'график _ 8' in fname or 'график _8' in fname or 'график 8' in fname:
        return '8'
    if 'график _ 9' in fname or 'график _9' in fname or 'график 9' in fname:
        return '9'
    if 'график_10' in fname:
        return '10'
    if 'график_11' in fname:
        return '11'
    if 'график_1,2' in fname:
        return '1,2'
    if 'график_2,3' in fname:
        return '2,3'
    logger.warning(f"Неизвестный паттерн транша: {fname}")
    return None


def tranche_match(cell_value, target_tranche: str) -> bool:
    cell_str = normalize_text(cell_value).lower()
    tranche = normalize_text(target_tranche).lower()
    exclusions = {'1': ['10', '11', '1,2'], '2': ['1,2', '2,3'], '3': ['2,3']}
    if tranche not in cell_str:
        return False
    excl = exclusions.get(tranche, [])
    if any(bad in cell_str for bad in excl):
        return False
    return True


def find_fio_row_in_col_c(ws, fio: str, logger_prefix: str = "") -> int | None:
    col_c_idx = 3
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_c_idx)
        if str(cell.value) == str(fio):
            logger.info(f"{logger_prefix} Найдено ФИО '{fio}' в строке {row} (ячейка C{row})")
            return row
    return None


def find_target_row(ws, fio_row: int, target_tranche: str, logger_prefix: str = "") -> int | None:
    col_c_idx = 3
    for row in range(fio_row + 1, ws.max_row + 1):
        cell_c = ws.cell(row=row, column=col_c_idx)
        if tranche_match(cell_c.value, target_tranche):
            logger.info(f"{logger_prefix} Найдена строка транша '{target_tranche}' в строке {row} (ячейка C{row})")
            return row
    return None


def collect_tranche_rows_below_fio(ws, fio_row: int, fio: str, logger_prefix: str = "") -> list[int]:
    rows = []
    col_c_idx = 3
    valid_tranches = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '1,2', '2,3']
    found_empty = False
    
    for row in range(fio_row + 1, ws.max_row + 1):
        value_c = ws.cell(row=row, column=col_c_idx).value
        value_c_str = str(value_c).strip() if value_c is not None else ""
        
        if value_c_str == "":
            found_empty = True
            continue
        
        is_tranche = any(tranche_match(value_c, t) for t in valid_tranches)
        
        if not is_tranche:
            if "транш" not in value_c_str.lower() or found_empty:
                break
            continue
        
        if found_empty:
            break
        
        rows.append(row)
    
    return rows


def has_minus_in_m_for_any_tranche(ws, fio_row: int, fio: str, logger_prefix: str = "") -> bool:
    tranche_rows = collect_tranche_rows_below_fio(ws, fio_row, fio, logger_prefix=logger_prefix)
    if not tranche_rows:
        return False
    
    for row in tranche_rows:
        value_m = ws.cell(row=row, column=13).value
        if cell_contains_minus(value_m):
            logger.warning(f"{logger_prefix} !!! ОБНАРУЖЕН МИНУС В M{row}: {value_m}")
            logger.warning(f"{logger_prefix} !!! РАЗНЕСЕНИЕ ПО ВСЕМ ТРАНШАМ ДЛЯ ЭТОГО ФИО БУДЕТ ОТМЕНЕНО")
            return True
    
    return False


def fill_izhs_ddu_values(ws, target_row, df_row, logger_prefix: str = "", write_sum=True):
    payment_total = df_row['общий ежемесячный платеж']
    payment_principal = df_row['платеж в счет погашения основного долга']
    payment_interest = df_row['платеж процентов']
    payment_date = df_row['дата']
    payment_sum = df_row['сумма']

    # Замена (без суммирования)
    ws.cell(row=target_row, column=14, value=payment_total)
    logger.info(f"{logger_prefix} Запись в N{target_row}: Общий Ежемесячный платеж = {payment_total}")
    ws.cell(row=target_row, column=17, value=payment_total)
    logger.info(f"{logger_prefix} Запись в Q{target_row}: Общий Ежемесячный платеж = {payment_total}")
    ws.cell(row=target_row, column=15, value=payment_principal)
    logger.info(f"{logger_prefix} Запись в O{target_row}: Платеж в счет погашения основного долга = {payment_principal}")
    ws.cell(row=target_row, column=18, value=payment_principal)
    logger.info(f"{logger_prefix} Запись в R{target_row}: Платеж в счет погашения основного долга = {payment_principal}")
    ws.cell(row=target_row, column=16, value=payment_interest)
    logger.info(f"{logger_prefix} Запись в P{target_row}: Платеж процентов = {payment_interest}")
    ws.cell(row=target_row, column=19, value=payment_interest)
    logger.info(f"{logger_prefix} Запись в S{target_row}: Платеж процентов = {payment_interest}")

    # Дата с накоплением через запятую
    cell_y = ws.cell(row=target_row, column=25)
    append_date_to_cell(cell_y, payment_date, logger_prefix)

    # Сумма с формулой и комментарием
    if write_sum:
        cell_u = ws.cell(row=target_row, column=21)
        add_to_cell_with_formula(cell_u, payment_sum, logger_prefix)
    else:
        logger.info(f"{logger_prefix} Запись в U{target_row} пропущена: сумма уже была внесена ранее для этого ФИО")


def process_izhs_ddu_row(df_row: pd.Series, logger_prefix: str, target_files, write_sum=True):
    vid = df_row['вид']
    fio = df_row['фио по реестру']
    file_graph = df_row['файл графика']

    if vid not in target_files:
        logger.warning(f"{logger_prefix} Неизвестный вид: {vid}")
        return False

    target_file = target_files[vid]
    if not target_file.exists():
        logger.warning(f"{logger_prefix} Файл не существует: {target_file}")
        return False

    try:
        wb = load_workbook(target_file)
        ws_name = wb.sheetnames[-1]
        ws = wb[ws_name]

        logger.info(f"{logger_prefix} Работаем с листом '{ws_name}' в {target_file}")
        logger.info(f"{logger_prefix} Исходные данные: вид = {vid}, ФИО = {fio}, файл графика = {file_graph}, write_sum = {write_sum}")

        fio_row = find_fio_row_in_col_c(ws, fio, logger_prefix=logger_prefix)
        if fio_row is None:
            logger.warning(f"{logger_prefix} ФИО '{fio}' не найдено в столбце C")
            wb.close()
            return False

        if has_minus_in_m_for_any_tranche(ws, fio_row, fio, logger_prefix=logger_prefix):
            wb.close()
            return 'debtor'

        tranche = get_tranche_from_filename(file_graph)
        if tranche is None:
            logger.warning(f"{logger_prefix} Транш не определён из '{file_graph}'")
            wb.close()
            return False

        logger.info(f"{logger_prefix} Транш: '{tranche}' из '{file_graph}'")

        target_row = find_target_row(ws, fio_row, tranche, logger_prefix=logger_prefix)
        if target_row is None:
            logger.warning(f"{logger_prefix} Строка транша '{tranche}' не найдена ниже строки {fio_row}")
            wb.close()
            return False

        logger.info(f"{logger_prefix} Целевая строка для заполнения: {target_row}")

        fill_izhs_ddu_values(ws, target_row, df_row, logger_prefix=logger_prefix, write_sum=write_sum)

        wb.save(target_file)
        wb.close()
        logger.info(f"{logger_prefix} Данные записаны и файл сохранён. Найденная строка: {target_row}")
        return True

    except Exception as e:
        logger.error(f"{logger_prefix} Ошибка обработки: {e}")
        return False


def main_izhs_ddu(df: pd.DataFrame, target_files: dict, txt_done_path: str) -> pd.DataFrame:
    relevant_rows = df[df['вид'].isin(['ИЖС', 'ДДУ'])].copy()
    logger.info(f"Обработка {len(relevant_rows)} строк ИЖС/ДДУ из {txt_done_path}")

    processed_sum_keys = set()
    blocked_keys = set()

    for idx, row in relevant_rows.iterrows():
        logger.info("")
        prefix = f"[Строка {idx + 2}]"

        vid = str(row.get('вид', '')).strip()
        fio = str(row.get('фио по реестру', '')).strip()
        sum_key = (vid, fio)

        logger.info(f"{prefix} Начало обработки строки")
        logger.info(f"{prefix} Ключ контроля суммы: {sum_key}")

        if sum_key in blocked_keys:
            logger.warning(f"{prefix} !!! РАЗНЕСЕНИЕ ПРОПУЩЕНО: для {sum_key} ранее уже установлен статус '{STATUS_DEBTOR}'")
            logger.info(f"{prefix} Обработка строки завершена")
            continue

        write_sum = sum_key not in processed_sum_keys
        logger.info(f"{prefix} Сумма будет записана: {write_sum}")

        result = process_izhs_ddu_row(row, prefix, target_files, write_sum=write_sum)

        if result == 'debtor':
            mask = (
                df['вид'].astype(str).str.strip().eq(vid) &
                df['фио по реестру'].astype(str).str.strip().eq(fio)
            )
            mark_rows_as_debtor(df, mask, logger_prefix=prefix)
            mark_debtor_in_excel_directly(txt_done_path, df, mask, STATUS_DEBTOR, logger_prefix=prefix)
            blocked_keys.add(sum_key)
            logger.warning(f"{prefix} !!! ВСЕ СТРОКИ ДЛЯ {sum_key} ПОМЕЧЕНЫ КАК '{STATUS_DEBTOR}' В txt_done.xlsx")
        elif result is True and write_sum:
            processed_sum_keys.add(sum_key)
            logger.info(f"{prefix} Ключ {sum_key} добавлен в processed_sum_keys")

        logger.info(f"{prefix} Обработка строки завершена")

    logger.info("")
    logger.info("-" * 100)
    logger.info(">>> ЭТАП 1 ЗАВЕРШЕН: ОБРАБОТКА ИЖС/ДДУ")
    logger.info("-" * 100)
    return df


# =========================
# ФУНКЦИИ ДЛЯ ЖСК
# =========================
def find_jsk_row(ws, fio: str, logger_prefix: str = "") -> int | None:
    col_fio_idx = 3
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_fio_idx)
        if str(cell.value) == str(fio):
            logger.info(f"{logger_prefix} Найдено ФИО '{fio}' в строке {row} (ячейка C{row})")
            return row
    return None


def fill_jsk_values(ws, target_row, df_row, logger_prefix: str = ""):
    payment_total = df_row['общий ежемесячный платеж']
    payment_principal = df_row['платеж в счет погашения основного долга']
    payment_interest = df_row['платеж процентов']
    payment_sum = df_row['сумма']
    payment_date = df_row['дата']

    # Замена (без суммирования)
    ws.cell(row=target_row, column=8, value=payment_total)
    logger.info(f"{logger_prefix} Запись в H{target_row}: Общий Ежемесячный платеж = {payment_total}")
    ws.cell(row=target_row, column=9, value=payment_principal)
    logger.info(f"{logger_prefix} Запись в I{target_row}: Платеж в счет погашения основного долга = {payment_principal}")
    ws.cell(row=target_row, column=12, value=payment_principal)
    logger.info(f"{logger_prefix} Запись в L{target_row}: Платеж в счет погашения основного долга = {payment_principal}")
    ws.cell(row=target_row, column=10, value=payment_interest)
    logger.info(f"{logger_prefix} Запись в J{target_row}: Платеж процентов = {payment_interest}")
    ws.cell(row=target_row, column=13, value=payment_interest)
    logger.info(f"{logger_prefix} Запись в M{target_row}: Платеж процентов = {payment_interest}")

    # Сумма с формулой и комментарием
    cell_k = ws.cell(row=target_row, column=11)
    add_to_cell_with_formula(cell_k, payment_sum, logger_prefix)

    # Дата с накоплением через запятую
    cell_s = ws.cell(row=target_row, column=19)
    append_date_to_cell(cell_s, payment_date, logger_prefix)


def process_jsk_row(df_row: pd.Series, logger_prefix: str, target_file: Path):
    vid = df_row['вид']
    fio = df_row['фио по реестру']

    if vid != 'ЖСК':
        logger.warning(f"{logger_prefix} Ожидался вид 'ЖСК', получено: {vid}")
        return False

    if not target_file.exists():
        logger.warning(f"{logger_prefix} Файл не существует: {target_file}")
        return False

    try:
        wb = load_workbook(target_file)
        ws_name = wb.sheetnames[-1]
        ws = wb[ws_name]

        logger.info(f"{logger_prefix} Работаем с листом '{ws_name}' в {target_file}")
        logger.info(f"{logger_prefix} Исходные данные: вид = {vid}, ФИО = {fio}")

        target_row = find_jsk_row(ws, fio, logger_prefix=logger_prefix)
        if target_row is None:
            logger.warning(f"{logger_prefix} ФИО '{fio}' не найдено в файле ЖСК")
            wb.close()
            return False

        value_g = ws.cell(row=target_row, column=7).value
        if cell_contains_minus(value_g):
            logger.warning(f"{logger_prefix} !!! ОБНАРУЖЕН МИНУС В G{target_row}: {value_g}")
            logger.warning(f"{logger_prefix} !!! РАЗНЕСЕНИЕ ПО ЖСК ДЛЯ ЭТОГО ФИО БУДЕТ ОТМЕНЕНО")
            wb.close()
            return 'debtor'

        logger.info(f"{logger_prefix} Целевая строка для заполнения: {target_row}")

        fill_jsk_values(ws, target_row, df_row, logger_prefix=logger_prefix)

        wb.save(target_file)
        wb.close()
        logger.info(f"{logger_prefix} Данные записаны и файл сохранён. Найденная строка: {target_row}")
        return True

    except Exception as e:
        logger.error(f"{logger_prefix} Ошибка обработки: {e}")
        return False


def main_jsk(df: pd.DataFrame, target_file: Path, txt_done_path: str) -> pd.DataFrame:
    relevant_rows = df[df['вид'] == 'ЖСК'].copy()
    logger.info(f"Обработка {len(relevant_rows)} строк ЖСК из {txt_done_path}")

    blocked_fio = set()

    for idx, row in relevant_rows.iterrows():
        logger.info("")
        prefix = f"[Строка {idx + 2}]"
        fio = str(row.get('фио по реестру', '')).strip()

        logger.info(f"{prefix} Начало обработки строки")
        logger.info(f"{prefix} ФИО из txt_done = {fio}")

        if fio in blocked_fio:
            logger.warning(f"{prefix} !!! РАЗНЕСЕНИЕ ПРОПУЩЕНО: для ФИО '{fio}' ранее уже установлен статус '{STATUS_DEBTOR}'")
            logger.info(f"{prefix} Обработка строки завершена")
            continue

        result = process_jsk_row(row, prefix, target_file)

        if result == 'debtor':
            mask = df['фио по реестру'].astype(str).str.strip().eq(fio) & df['вид'].astype(str).str.strip().eq('ЖСК')
            mark_rows_as_debtor(df, mask, logger_prefix=prefix)
            mark_debtor_in_excel_directly(txt_done_path, df, mask, STATUS_DEBTOR, logger_prefix=prefix)
            blocked_fio.add(fio)
            logger.warning(f"{prefix} !!! ВСЕ СТРОКИ ЖСК ДЛЯ '{fio}' ПОМЕЧЕНЫ КАК '{STATUS_DEBTOR}' В txt_done.xlsx")

        logger.info(f"{prefix} Обработка строки завершена")

    logger.info("")
    logger.info("-" * 100)
    logger.info(">>> ЭТАП 2 ЗАВЕРШЕН: ОБРАБОТКА ЖСК")
    logger.info("-" * 100)
    return df


# =========================
# ФУНКЦИИ ДЛЯ ОСНОВНОГО РЕЕСТРА
# =========================
def find_fio_row(ws, fio: str, logger_prefix: str = "") -> int | None:
    col_fio_idx = 4
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_fio_idx)
        if str(cell.value) == str(fio):
            logger.info(f"{logger_prefix} Найдено ФИО '{fio}' в строке {row} (ячейка D{row})")
            return row
    return None


def find_month_year_col(ws, month_num, year, logger_prefix: str = "") -> int | None:
    try:
        month_name = MONTH_NAMES_RU.get(int(month_num))
        if not month_name:
            logger.warning(f"{logger_prefix} Неизвестный месяц: {month_num}")
            return None
        month_name = month_name.lower()
        year_str = str(int(year))
    except Exception:
        logger.warning(f"{logger_prefix} Некорректные месяц/год: month={month_num}, year={year}")
        return None

    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        header_str = str(header_value).lower() if header_value is not None else ""
        if month_name in header_str and year_str in header_str:
            logger.info(f"{logger_prefix} Найден столбец '{header_value}' в колонке {col}")
            return col
    return None


def fill_main_registry_sum(ws, fio_row, month_col, df_row, logger_prefix: str = ""):
    payment_sum = df_row['сумма']
    cell = ws.cell(row=fio_row, column=month_col)
    add_to_cell_with_formula(cell, payment_sum, logger_prefix)


def process_main_registry_row(df_row: pd.Series, logger_prefix: str, target_file: Path):
    fio = df_row['фио по реестру']
    month_num = df_row['месяц платежа']
    year = df_row['год платежа']

    if not target_file.exists():
        logger.warning(f"{logger_prefix} Файл не существует: {target_file}")
        return False

    try:
        wb = load_workbook(target_file)
        ws_name = wb.sheetnames[0]
        ws = wb[ws_name]

        logger.info(f"{logger_prefix} Работаем с листом '{ws_name}' в {target_file}")
        logger.info(f"{logger_prefix} Исходные данные: ФИО = {fio}, месяц платежа = {month_num}, год платежа = {year}, сумма = {df_row['сумма']}")

        fio_row = find_fio_row(ws, fio, logger_prefix=logger_prefix)
        if fio_row is None:
            logger.warning(f"{logger_prefix} ФИО '{fio}' не найдено")
            wb.close()
            return False

        month_col = find_month_year_col(ws, month_num, year, logger_prefix=logger_prefix)
        if month_col is None:
            logger.warning(f"{logger_prefix} Не найден столбец для месяца '{month_num}' и года '{year}'")
            wb.close()
            return False

        logger.info(f"{logger_prefix} Целевая ячейка: строка {fio_row}, столбец {month_col}")

        fill_main_registry_sum(ws, fio_row, month_col, df_row, logger_prefix=logger_prefix)

        wb.save(target_file)
        wb.close()
        logger.info(f"{logger_prefix} Данные записаны и файл сохранён")
        return True

    except Exception as e:
        logger.error(f"{logger_prefix} Ошибка обработки: {e}")
        return False


def main_main_registry(df: pd.DataFrame, target_file: Path, txt_done_path: str) -> pd.DataFrame:
    filtered_df = df[~df[STATUS_COL_NAME].apply(is_debtor_status)].copy()
    logger.info(f"Обработка строк для основного реестра из {txt_done_path}")
    logger.info(f"После исключения строк со статусом '{STATUS_DEBTOR}' осталось строк: {len(filtered_df)}")

    processed_fio = set()

    for idx, row in filtered_df.iterrows():
        logger.info("")
        prefix = f"[Строка {idx + 2}]"
        fio = str(row.get('фио по реестру', '')).strip()

        logger.info(f"{prefix} Начало обработки строки")

        if fio in processed_fio:
            logger.info(f"{prefix} ФИО '{fio}' уже обработано ранее, строка пропущена")
            logger.info(f"{prefix} Обработка строки завершена")
            continue

        process_main_registry_row(row, prefix, target_file)
        processed_fio.add(fio)

        logger.info(f"{prefix} ФИО '{fio}' добавлено в processed_fio")
        logger.info(f"{prefix} Обработка строки завершена")

    logger.info("")
    logger.info("-" * 100)
    logger.info(">>> ЭТАП 3 ЗАВЕРШЕН: ОБРАБОТКА ОСНОВНОГО РЕЕСТРА")
    logger.info("-" * 100)
    return df


# =========================
# ОСНОВНАЯ ФУНКЦИЯ ОБРАБОТКИ
# =========================
def run_full_processing(txt_done_path: str, paths: dict, create_backup_flag: bool, progress_callback=None):
    """
    paths = {
        "ИЖС": Path,
        "ДДУ": Path,
        "ЖСК": Path,
        "основной": Path
    }
    """
    logger.info("=" * 100)
    logger.info(f">> СТАРТ {APP_NAME}")
    logger.info("=" * 100)
    logger.info(f"Исходный файл: {txt_done_path}")
    logger.info(f"ИЖС-реестр: {paths['ИЖС']}")
    logger.info(f"ДДУ-реестр: {paths['ДДУ']}")
    logger.info(f"ЖСК-реестр: {paths['ЖСК']}")
    logger.info(f"Основной реестр: {paths['основной']}")
    logger.info(f"Создание резервных копий: {'ДА' if create_backup_flag else 'НЕТ'}")
    logger.info("=" * 100)

    # Создание резервных копий
    if create_backup_flag:
        backup_sources = {
            "txt_done": Path(txt_done_path),
            "ИЖС": paths["ИЖС"],
            "ДДУ": paths["ДДУ"],
            "ЖСК": paths["ЖСК"],
            "основной": paths["основной"]
        }
        backup_dir = create_backup(backup_sources)
        if backup_dir is None:
            logger.error("Не удалось создать резервные копии. Обработка прервана.")
            if progress_callback:
                progress_callback(0, "ошибка")
            return

    if progress_callback:
        progress_callback(0.1, "Загрузка данных...")

    df = load_txt_done_df(txt_done_path)
    logger.info(f"Загружено строк из txt_done.xlsx: {len(df)}")

    if progress_callback:
        progress_callback(0.2, "Обработка ИЖС/ДДУ...")

    target_files_izhs_ddu = {
        "ИЖС": paths["ИЖС"],
        "ДДУ": paths["ДДУ"]
    }
    df = main_izhs_ddu(df, target_files_izhs_ddu, txt_done_path)

    if progress_callback:
        progress_callback(0.5, "Обработка ЖСК...")

    df = main_jsk(df, paths["ЖСК"], txt_done_path)

    if progress_callback:
        progress_callback(0.7, "Обработка основного реестра...")

    df = main_main_registry(df, paths["основной"], txt_done_path)

    if progress_callback:
        progress_callback(0.95, "Сохранение результатов...")

    logger.info("=" * 100)
    logger.info(">> ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО")
    logger.info("=" * 100)

    if progress_callback:
        progress_callback(1.0, "Готово!")


# =========================
# ГЛАВНОЕ ОКНО ПРИЛОЖЕНИЯ
# =========================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.log_queue = queue.Queue()
        setup_logger(self.log_queue)

        self.title(APP_NAME)
        self.geometry("1000x900")
        self.minsize(900, 750)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.worker_thread = None
        self.create_backup_var = ctk.BooleanVar(value=True)

        # Настройка растяжения строк
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=0)
        self.grid_rowconfigure(3, weight=0)
        self.grid_rowconfigure(4, weight=0)
        self.grid_rowconfigure(5, weight=0)
        self.grid_rowconfigure(6, weight=1)

        self.build_ui()
        self.after(100, self.poll_log_queue)

    def build_ui(self):
        # Верхняя панель с заголовком и логотипом
        self.top_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.top_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        self.top_frame.grid_columnconfigure(0, weight=1)

        self.header_label = ctk.CTkLabel(
            self.top_frame,
            text=APP_NAME,
            font=ctk.CTkFont(size=28, weight="bold")
        )
        self.header_label.grid(row=0, column=0, sticky="w")

        logo_path = get_resource_path(LOGO_FILENAME)
        try:
            if os.path.exists(logo_path):
                logo_image = Image.open(logo_path)
                new_height = 70
                original_width, original_height = logo_image.size
                new_width = int(original_width * (new_height / original_height))
                logo_ctk = ctk.CTkImage(light_image=logo_image, dark_image=logo_image, size=(new_width, new_height))
                self.logo_label = ctk.CTkLabel(self.top_frame, image=logo_ctk, text="")
                self.logo_label.grid(row=0, column=1, padx=(15, 0), sticky="e")
        except Exception as e:
            pass

        # Рамка с полями ввода
        self.input_frame = ctk.CTkFrame(self)
        self.input_frame.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")
        self.input_frame.grid_columnconfigure(1, weight=1)

        row = 0

        self.lbl_source = ctk.CTkLabel(self.input_frame, text="Исходный файл (txt_done.xlsx):")
        self.lbl_source.grid(row=row, column=0, padx=12, pady=(12, 6), sticky="w")

        self.entry_source = ctk.CTkEntry(self.input_frame)
        self.entry_source.grid(row=row, column=1, padx=12, pady=(12, 6), sticky="ew")

        self.btn_source = ctk.CTkButton(self.input_frame, text="Обзор", width=100, command=self.choose_source)
        self.btn_source.grid(row=row, column=2, padx=12, pady=(12, 6))

        row += 1
        self.lbl_output = ctk.CTkLabel(self.input_frame, text="Финальный файл (сохраняется в ту же папку):")
        self.lbl_output.grid(row=row, column=0, padx=12, pady=6, sticky="w")

        self.entry_output = ctk.CTkEntry(self.input_frame, state="readonly")
        self.entry_output.grid(row=row, column=1, padx=12, pady=6, sticky="ew")

        self.btn_open_folder = ctk.CTkButton(self.input_frame, text="Открыть папку", width=100, command=self.open_output_folder)
        self.btn_open_folder.grid(row=row, column=2, padx=12, pady=6)

        row += 1
        self.sep1 = ctk.CTkLabel(self.input_frame, text="", height=10)
        self.sep1.grid(row=row, column=0, columnspan=3, sticky="ew")

        row += 1
        self.lbl_registry_header = ctk.CTkLabel(
            self.input_frame,
            text="Целевые реестры (можно изменить при необходимости):",
            font=ctk.CTkFont(weight="bold")
        )
        self.lbl_registry_header.grid(row=row, column=0, columnspan=3, padx=12, pady=(10, 5), sticky="w")

        row += 1
        self.lbl_izhs = ctk.CTkLabel(self.input_frame, text="ИЖС-реестр:", width=120, anchor="w")
        self.lbl_izhs.grid(row=row, column=0, padx=12, pady=6, sticky="w")

        self.entry_izhs = ctk.CTkEntry(self.input_frame)
        self.entry_izhs.grid(row=row, column=1, padx=12, pady=6, sticky="ew")
        self.entry_izhs.insert(0, str(DEFAULT_PATHS["ИЖС"]))

        self.btn_izhs = ctk.CTkButton(self.input_frame, text="Обзор", width=100, command=lambda: self.browse_file(self.entry_izhs))
        self.btn_izhs.grid(row=row, column=2, padx=12, pady=6)

        row += 1
        self.lbl_ddu = ctk.CTkLabel(self.input_frame, text="ДДУ-реестр:", width=120, anchor="w")
        self.lbl_ddu.grid(row=row, column=0, padx=12, pady=6, sticky="w")

        self.entry_ddu = ctk.CTkEntry(self.input_frame)
        self.entry_ddu.grid(row=row, column=1, padx=12, pady=6, sticky="ew")
        self.entry_ddu.insert(0, str(DEFAULT_PATHS["ДДУ"]))

        self.btn_ddu = ctk.CTkButton(self.input_frame, text="Обзор", width=100, command=lambda: self.browse_file(self.entry_ddu))
        self.btn_ddu.grid(row=row, column=2, padx=12, pady=6)

        row += 1
        self.lbl_jsk = ctk.CTkLabel(self.input_frame, text="ЖСК-реестр:", width=120, anchor="w")
        self.lbl_jsk.grid(row=row, column=0, padx=12, pady=6, sticky="w")

        self.entry_jsk = ctk.CTkEntry(self.input_frame)
        self.entry_jsk.grid(row=row, column=1, padx=12, pady=6, sticky="ew")
        self.entry_jsk.insert(0, str(DEFAULT_PATHS["ЖСК"]))

        self.btn_jsk = ctk.CTkButton(self.input_frame, text="Обзор", width=100, command=lambda: self.browse_file(self.entry_jsk))
        self.btn_jsk.grid(row=row, column=2, padx=12, pady=6)

        row += 1
        self.lbl_main = ctk.CTkLabel(self.input_frame, text="Основной реестр:", width=120, anchor="w")
        self.lbl_main.grid(row=row, column=0, padx=12, pady=6, sticky="w")

        self.entry_main = ctk.CTkEntry(self.input_frame)
        self.entry_main.grid(row=row, column=1, padx=12, pady=6, sticky="ew")
        self.entry_main.insert(0, str(DEFAULT_PATHS["основной"]))

        self.btn_main = ctk.CTkButton(self.input_frame, text="Обзор", width=100, command=lambda: self.browse_file(self.entry_main))
        self.btn_main.grid(row=row, column=2, padx=12, pady=6)

        # Рамка с настройками и кнопками
        self.action_frame = ctk.CTkFrame(self)
        self.action_frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        self.action_frame.grid_columnconfigure(0, weight=0)
        self.action_frame.grid_columnconfigure(1, weight=0)
        self.action_frame.grid_columnconfigure(2, weight=0)
        self.action_frame.grid_columnconfigure(3, weight=1)

        self.btn_run = ctk.CTkButton(self.action_frame, text="ЗАПУСТИТЬ РАЗНЕСЕНИЕ", command=self.start_processing, width=200, height=40)
        self.btn_run.grid(row=0, column=0, padx=12, pady=12, sticky="w")

        self.btn_restore = ctk.CTkButton(self.action_frame, text="Восстановить из резервной копии", command=self.restore_from_backup, width=220, height=40)
        self.btn_restore.grid(row=0, column=1, padx=12, pady=12, sticky="w")

        self.btn_clear = ctk.CTkButton(self.action_frame, text="Очистить лог", command=self.clear_log, width=140)
        self.btn_clear.grid(row=0, column=2, padx=12, pady=12, sticky="w")

        self.status_label = ctk.CTkLabel(self.action_frame, text="Статус: готово", anchor="w")
        self.status_label.grid(row=0, column=3, padx=12, pady=12, sticky="ew")

        # Галочка резервного копирования
        self.backup_frame = ctk.CTkFrame(self)
        self.backup_frame.grid(row=3, column=0, padx=20, pady=5, sticky="ew")
        self.backup_frame.grid_columnconfigure(0, weight=0)

        self.backup_checkbox = ctk.CTkCheckBox(
            self.backup_frame,
            text="Создать резервные копии перед запуском",
            variable=self.create_backup_var
        )
        self.backup_checkbox.grid(row=0, column=0, padx=12, pady=5, sticky="w")

        # Прогресс-бар
        self.progress = ctk.CTkProgressBar(self)
        self.progress.grid(row=4, column=0, padx=20, pady=(0, 10), sticky="ew")
        self.progress.set(0)

        # Информационная строка
        self.info_label = ctk.CTkLabel(
            self,
            text="Результат сохраняется в исходный файл (txt_done.xlsx) в ту же папку",
            justify="left"
        )
        self.info_label.grid(row=5, column=0, padx=20, pady=(0, 10), sticky="w")

        # Лог-окно
        self.log_box = ctk.CTkTextbox(self, wrap="word")
        self.log_box.grid(row=6, column=0, padx=20, pady=(0, 20), sticky="nsew")
        self.log_box.configure(state="disabled")

        self.log_box.tag_config("info", foreground="white")
        self.log_box.tag_config("warning", foreground="yellow")
        self.log_box.tag_config("error", foreground="orange")
        self.log_box.tag_config("critical", foreground="red")

    def browse_file(self, entry_widget):
        path = filedialog.askopenfilename(
            title="Выберите файл",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xls"), ("All files", "*.*")]
        )
        if path:
            entry_widget.delete(0, "end")
            entry_widget.insert(0, path)

    def choose_source(self):
        path = filedialog.askopenfilename(
            title="Выберите txt_done.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.entry_source.delete(0, "end")
            self.entry_source.insert(0, path)
            self.update_output_path()

    def update_output_path(self):
        source_path = self.entry_source.get().strip()
        if source_path:
            output_path = str(Path(source_path).parent / "txt_done.xlsx")
        else:
            output_path = ""

        self.entry_output.configure(state="normal")
        self.entry_output.delete(0, "end")
        self.entry_output.insert(0, output_path)
        self.entry_output.configure(state="readonly")

    def open_output_folder(self):
        output_path = self.entry_output.get().strip()
        if output_path:
            folder = str(Path(output_path).parent)
            if os.path.exists(folder):
                os.startfile(folder)
            else:
                messagebox.showwarning("Ошибка", f"Папка не существует: {folder}")

    def clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def append_log(self, text):
        self.log_box.configure(state="normal")

        if "| ERROR |" in text or "ОШИБКА" in text.upper():
            tag = "error"
        elif "| WARNING |" in text or "WARNING" in text.upper():
            tag = "warning"
        elif "| CRITICAL |" in text or "КРИТИЧЕСКАЯ" in text.upper():
            tag = "critical"
        else:
            tag = "info"

        self.log_box.insert("end", text + "\n", tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def update_progress(self, value, status_text=None):
        self.progress.set(value)
        if status_text:
            self.status_label.configure(text=f"Статус: {status_text}")

    def poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.append_log(msg)

                if "СТАРТ" in msg:
                    self.update_progress(0.05, "обработка...")
                elif "Загружено строк из txt_done.xlsx" in msg:
                    self.update_progress(0.15, "загрузка данных...")
                elif ">>> ЭТАП 1 ЗАВЕРШЕН" in msg:
                    self.update_progress(0.45, "обработка ИЖС/ДДУ...")
                elif ">>> ЭТАП 2 ЗАВЕРШЕН" in msg:
                    self.update_progress(0.65, "обработка ЖСК...")
                elif ">>> ЭТАП 3 ЗАВЕРШЕН" in msg:
                    self.update_progress(0.85, "обработка основного реестра...")
                elif "ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО" in msg:
                    self.update_progress(1.0, "готово!")
                elif "КРИТИЧЕСКАЯ ОШИБКА" in msg.upper():
                    self.update_progress(0, "ошибка")
        except queue.Empty:
            pass

        self.after(100, self.poll_log_queue)

    def set_controls_state(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        self.btn_run.configure(state=state)
        self.btn_restore.configure(state=state)
        self.btn_clear.configure(state=state)
        self.btn_source.configure(state=state)
        self.btn_izhs.configure(state=state)
        self.btn_ddu.configure(state=state)
        self.btn_jsk.configure(state=state)
        self.btn_main.configure(state=state)
        self.entry_source.configure(state=state)
        self.entry_izhs.configure(state=state)
        self.entry_ddu.configure(state=state)
        self.entry_jsk.configure(state=state)
        self.entry_main.configure(state=state)
        self.backup_checkbox.configure(state=state)
        self.btn_open_folder.configure(state="normal")

    def confirm_and_run(self):
        result = messagebox.askyesno(
            "Подтверждение запуска",
            "Перед продолжением убедитесь, что все реестры и txt_done.xlsx закрыты в Excel.\n\n"
            "Нажмите 'Да', чтобы продолжить, или 'Нет' для отмены.",
            icon="warning"
        )
        if result:
            self.start_processing_real()

    def start_processing(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return

        source_path = self.entry_source.get().strip()
        if not source_path:
            self.append_log("Ошибка: не выбран исходный файл txt_done.xlsx")
            self.status_label.configure(text="Статус: ошибка")
            return

        if not Path(source_path).exists():
            self.append_log(f"Ошибка: файл {source_path} не существует")
            self.status_label.configure(text="Статус: ошибка")
            return

        self.confirm_and_run()

    def start_processing_real(self):
        source_path = self.entry_source.get().strip()

        paths = {
            "ИЖС": Path(self.entry_izhs.get().strip()),
            "ДДУ": Path(self.entry_ddu.get().strip()),
            "ЖСК": Path(self.entry_jsk.get().strip()),
            "основной": Path(self.entry_main.get().strip())
        }

        all_exist = True
        for key, p in paths.items():
            if not p.exists():
                self.append_log(f"Ошибка: файл {key} не существует: {p}")
                all_exist = False
        if not all_exist:
            self.status_label.configure(text="Статус: ошибка")
            return

        self.progress.set(0)
        self.status_label.configure(text="Статус: обработка...")
        self.set_controls_state(False)
        self.append_log("=" * 100)
        self.append_log(">> ЗАПУСК ОБРАБОТКИ")
        self.append_log("=" * 100)

        create_backup_flag = self.create_backup_var.get()

        def progress_callback(value, status_text=None):
            self.after(0, lambda: self.update_progress(value, status_text))

        self.worker_thread = threading.Thread(
            target=self.worker_run,
            args=(source_path, paths, create_backup_flag, progress_callback),
            daemon=True
        )
        self.worker_thread.start()

    def worker_run(self, source_path: str, paths: dict, create_backup_flag: bool, progress_callback):
        try:
            run_full_processing(source_path, paths, create_backup_flag, progress_callback)
        except Exception as e:
            logger.error("=" * 100)
            logger.error(">> КРИТИЧЕСКАЯ ОШИБКА ВЫПОЛНЕНИЯ")
            logger.error("=" * 100)
            logger.error(str(e))
            logger.error(traceback.format_exc())
            self.after(0, lambda: self.update_progress(0, "ошибка"))
        finally:
            self.after(0, lambda: self.set_controls_state(True))

    def restore_from_backup(self):
        """Восстановление файлов из последней резервной копии"""
        backups = get_all_backups()

        if not backups:
            messagebox.showwarning(
                "Нет резервных копий",
                "Резервные копии не найдены.\n\n"
                "Чтобы создать резервную копию, поставьте галочку 'Создать резервные копии перед запуском' и выполните обработку."
            )
            return

        # Если несколько копий, показываем выбор
        if len(backups) == 1:
            selected_backup = backups[0]
        else:
            # Показываем список всех копий в простом диалоге
            backup_names = "\n".join([f"{i+1}. {b.name}" for i, b in enumerate(backups)])
            result = messagebox.askquestion(
                "Выбор резервной копии",
                f"Доступные резервные копии:\n\n{backup_names}\n\n"
                f"Введите номер копии (1-{len(backups)}) и нажмите 'Да'.\n\n"
                f"Или нажмите 'Нет' для отмены."
            )
            if result == "yes":
                try:
                    num = simpledialog.askinteger("Номер копии", f"Введите номер (1-{len(backups)}):", minvalue=1, maxvalue=len(backups))
                    if num is None:
                        return
                    selected_backup = backups[num - 1]
                except:
                    messagebox.showerror("Ошибка", "Не удалось выбрать копию")
                    return
            else:
                return

        # Проверяем существование целевых файлов
        source_path = self.entry_source.get().strip()
        if not source_path:
            messagebox.showerror("Ошибка", "Не выбран исходный файл txt_done.xlsx")
            return

        target_files = {
            "txt_done": Path(source_path),
            "ИЖС": Path(self.entry_izhs.get().strip()),
            "ДДУ": Path(self.entry_ddu.get().strip()),
            "ЖСК": Path(self.entry_jsk.get().strip()),
            "основной": Path(self.entry_main.get().strip())
        }

        # Предупреждение о закрытии файлов
        confirm = messagebox.askyesno(
            "Подтверждение восстановления",
            "ВНИМАНИЕ!\n\n"
            "Перед восстановлением убедитесь, что все файлы закрыты в Excel.\n\n"
            "Будут восстановлены файлы:\n"
            f"  - {target_files['txt_done'].name}\n"
            f"  - {target_files['ИЖС'].name}\n"
            f"  - {target_files['ДДУ'].name}\n"
            f"  - {target_files['ЖСК'].name}\n"
            f"  - {target_files['основной'].name}\n\n"
            "Восстановить из копии?\n\n"
            f"Копия: {selected_backup.name}",
            icon="warning"
        )

        if not confirm:
            return

        self.append_log("=" * 100)
        self.append_log(">> ЗАПУСК ВОССТАНОВЛЕНИЯ ИЗ РЕЗЕРВНОЙ КОПИИ")
        self.append_log("=" * 100)

        success = restore_from_backup(selected_backup, target_files)

        if success:
            messagebox.showinfo("Восстановление завершено", "Файлы успешно восстановлены из резервной копии!")
            self.append_log("Восстановление завершено успешно!")
        else:
            messagebox.showerror("Ошибка восстановления", "При восстановлении произошли ошибки. Проверьте лог.")
            self.append_log("Восстановление завершено с ошибками!")

        self.append_log("=" * 100)


# =========================
# ТОЧКА ВХОДА
# =========================
if __name__ == "__main__":
    app = App()
    app.mainloop()