import os
import sys
import queue
import shutil
import logging
import traceback
import threading
import re
import pythoncom
from pathlib import Path
from datetime import datetime
from tkinter import messagebox, filedialog

import pandas as pd
import customtkinter as ctk
from PIL import Image
from openpyxl import load_workbook
import win32com.client
from win32com.client import constants

APP_NAME = "Заливка_графиков"
LOGO_FILENAME = "logo-sia.png"

EXE_DIR = Path(sys.executable).resolve().parent if getattr(sys, 'frozen', False) else Path.cwd()

# =========================
# НАСТРОЙКИ ЛОГИРОВАНИЯ
# =========================
LOGS_BASE_DIR = EXE_DIR / "Логи"
TODAY = datetime.now().strftime("%Y-%m-%d")
TIME_NOW = datetime.now().strftime("%H-%M-%S")
LOGS_TODAY_DIR = LOGS_BASE_DIR / f"Логи_{TODAY}"
LOGS_TODAY_DIR.mkdir(parents=True, exist_ok=True)

LOG_PATH = LOGS_TODAY_DIR / f"{APP_NAME}_{TODAY}_{TIME_NOW}.log"

BACKUP_BASE_DIR = EXE_DIR / "резервные_копии"
BACKUP_BASE_DIR.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger(APP_NAME)

STATUS_COL_NAME = "Статус"
FILE_COL_NAME = "файл графика"


def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


class QueueLogHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record):
        try:
            msg = self.format(record)
            self.log_queue.put(msg)
        except Exception:
            pass


def setup_logger(log_queue=None):
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    if log_queue is not None:
        queue_handler = QueueLogHandler(log_queue)
        queue_handler.setFormatter(formatter)
        logger.addHandler(queue_handler)


# =========================
# ЗАГРУЗКА TXT_DONE С ГИПЕРССЫЛКАМИ
# =========================
def load_txt_done_with_hyperlinks(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip() if val else "")

    file_col_idx = None
    for i, h in enumerate(headers):
        if h.lower() == FILE_COL_NAME.lower():
            file_col_idx = i + 1
            break
    if file_col_idx is None:
        raise ValueError(f"Колонка '{FILE_COL_NAME}' не найдена")

    data = []
    hyperlinks = {}
    for row in range(2, ws.max_row + 1):
        row_data = []
        row_hyperlinks = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.hyperlink:
                link = cell.hyperlink.target
                if link:
                    link = link.replace('%20', ' ')
                row_hyperlinks[col - 1] = link
                row_data.append(cell.value)
            else:
                row_data.append(cell.value)
        data.append(row_data)
        if row_hyperlinks:
            hyperlinks[row - 2] = row_hyperlinks
    df = pd.DataFrame(data, columns=headers)
    wb.close()
    return df, hyperlinks, file_col_idx - 1


# =========================
# РЕЗЕРВНОЕ КОПИРОВАНИЕ
# =========================
def create_backup_for_files(files_to_backup, txt_done_path):
    """Создаёт резервные копии переданных файлов и txt_done.xlsx"""
    backup_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_dir = BACKUP_BASE_DIR / backup_time
    backup_dir.mkdir(parents=True, exist_ok=True)
    logger.info(">>> СОЗДАНИЕ РЕЗЕРВНЫХ КОПИЙ")
    success_count = 0
    for path in files_to_backup:
        if os.path.exists(path):
            try:
                shutil.copy2(path, backup_dir / Path(path).name)
                success_count += 1
            except Exception as e:
                logger.error(f"Ошибка копирования {Path(path).name}: {e}")
                return None
        else:
            logger.warning(f"Файл не найден: {path}")
    if txt_done_path.exists():
        shutil.copy2(txt_done_path, backup_dir / "txt_done_backup.xlsx")
    logger.info(f"Создано копий: {success_count}, папка: {backup_dir}")
    return backup_dir


def get_all_backups():
    backups = [b for b in BACKUP_BASE_DIR.iterdir() if b.is_dir()]
    backups.sort(key=lambda x: x.stat().st_ctime, reverse=True)
    return backups


def restore_from_backup(backup_dir, txt_done_path, graph_paths):
    logger.info(">>> ВОССТАНОВЛЕНИЕ ИЗ РЕЗЕРВНОЙ КОПИИ")
    success = True
    txt_backup = backup_dir / "txt_done_backup.xlsx"
    if txt_backup.exists():
        shutil.copy2(txt_backup, txt_done_path)
    for graph_path in graph_paths:
        backup_file = backup_dir / Path(graph_path).name
        if backup_file.exists():
            shutil.copy2(backup_file, graph_path)
        else:
            logger.warning(f"{Path(graph_path).name} не найден в бэкапе")
            success = False
    return success


# =========================
# РАБОТА С XLS ЧЕРЕЗ WIN32COM
# =========================
def normalize_value(value):
    """
    Нормализует значение для сравнения.
    Если значение можно преобразовать в число, округляет до 2 знаков и возвращает float.
    Иначе возвращает строку без лишних пробелов и .0 в конце.
    """
    if value is None or pd.isna(value):
        return None
    # Пробуем преобразовать в число
    try:
        num = float(value)
        # Округляем до 2 знаков для денежных сумм
        rounded = round(num, 2)
        return rounded
    except (ValueError, TypeError):
        # Если не число — возвращаем строку
        s = str(value).strip()
        if not s:
            return None
        # Убираем .0 в конце, если это целое число в строке
        s = re.sub(r"\.0$", "", s)
        return s


def RGB(r, g, b):
    # Excel использует BGR
    return (b << 16) + (g << 8) + r


def find_last_yellow_row_excel(ws):
    last_yellow_row = None
    try:
        last_cell = ws.Cells.SpecialCells(11)
        max_row = last_cell.Row if last_cell else ws.UsedRange.Rows.Count
    except:
        max_row = ws.UsedRange.Rows.Count
    if max_row < 10:
        max_row = 200

    for row in range(1, max_row + 1):
        is_yellow = True
        for col in range(4, 9):
            try:
                cell = ws.Cells(row, col)
                color_index = cell.Interior.ColorIndex
                if color_index not in [6, 44, 46, 27, 28]:
                    is_yellow = False
                    break
            except:
                is_yellow = False
                break
        if is_yellow:
            last_yellow_row = row
    return last_yellow_row


def paint_row_in_xls(xls_path, row_num):
    """Закрашивает указанную строку в XLS файле через win32com"""
    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(xls_path))
        if wb.Sheets.Count >= 2:
            ws = wb.Sheets(2)
        else:
            ws = wb.Sheets(1)
        last_col = ws.UsedRange.Columns.Count
        for col in range(1, last_col + 1):
            ws.Cells(row_num, col).Interior.Color = RGB(255, 255, 0)
        wb.Save()
        return True
    except Exception as e:
        logger.error(f"Ошибка при закрашивании {xls_path}: {e}")
        return False
    finally:
        try:
            if wb:
                wb.Close(SaveChanges=False)
            if excel:
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()


# =========================
# ОСНОВНАЯ ЛОГИКА
# =========================
def collect_rows_to_paint(txt_done_path, progress_callback=None):
    """
    Собирает данные о строках, которые нужно закрасить.
    Возвращает список словарей с ключами:
        txt_row, file_path, file_name, row_to_paint, month, year
    и список предупреждений (для лога).
    """
    logger.info("=" * 80)
    logger.info(f">> СТАРТ {APP_NAME} (сбор данных)")
    logger.info(f"Файл: {txt_done_path}")
    logger.info("=" * 80)

    if not txt_done_path.exists():
        raise FileNotFoundError(f"Файл не найден: {txt_done_path}")

    if progress_callback:
        progress_callback(0.05, "Загрузка...")

    df, hyperlinks, file_col_idx = load_txt_done_with_hyperlinks(txt_done_path)
    logger.info(f"Загружено строк данных: {len(df)}")

    # Добавляем колонку Статус, если нет
    if STATUS_COL_NAME not in df.columns:
        df[STATUS_COL_NAME] = ""
        cols = df.columns.tolist()
        cols.remove(STATUS_COL_NAME)
        cols.append(STATUS_COL_NAME)
        df = df[cols]

    # Индексы колонок
    col_map = {}
    for col in ['дата', 'договор', 'месяц платежа', 'год платежа',
                'общий ежемесячный платеж', 'платеж в счет погашения основного долга',
                'платеж процентов', FILE_COL_NAME]:
        for i, c in enumerate(df.columns):
            if str(c).strip().lower() == col.lower():
                col_map[col] = i
                break
    for col in ['дата', 'договор', 'месяц платежа', 'год платежа',
                'общий ежемесячный платеж', 'платеж в счет погашения основного долга',
                'платеж процентов', FILE_COL_NAME]:
        if col not in col_map:
            raise ValueError(f"Колонка '{col}' не найдена")

    status_idx = None
    for i, col in enumerate(df.columns):
        if str(col).strip() == STATUS_COL_NAME:
            status_idx = i
            break
    if status_idx is None:
        df[STATUS_COL_NAME] = ""
        status_idx = len(df.columns) - 1
    logger.info(f"Колонка '{STATUS_COL_NAME}' индекс: {status_idx}")

    rows_to_paint = []
    warnings = []

    logger.info("")
    logger.info(">>> СБОР ДАННЫХ ДЛЯ ЗАКРАШИВАНИЯ")
    total_rows = len(df)
    for idx, row in df.iterrows():
        if progress_callback:
            progress_val = 0.05 + ((idx + 1) / total_rows) * 0.4
            progress_callback(progress_val, f"сбор {idx+1}/{total_rows}")

        # Проверяем статус
        status_val = str(row.iloc[status_idx]).strip() if status_idx < len(row) else ""
        if status_val and status_val.lower() not in ['nan', 'none', '']:
            logger.info(f"Строка {idx+2}: статус '{status_val}' – пропуск")
            continue

        # Получаем путь к файлу графика
        file_path = None
        if idx in hyperlinks and file_col_idx in hyperlinks[idx]:
            file_path = hyperlinks[idx][file_col_idx]
        if not file_path or not str(file_path).strip():
            file_path = row.iloc[file_col_idx] if file_col_idx < len(row) else None
        if not file_path or pd.isna(file_path) or not str(file_path).strip():
            warnings.append(f"Строка {idx+2}: путь к файлу пуст")
            continue
        file_path = str(file_path).strip()
        if file_path.lower().startswith("file:///"):
            file_path = file_path[8:].replace("/", "\\")
        if not os.path.exists(file_path):
            warnings.append(f"Строка {idx+2}: файл не найден: {file_path}")
            continue
        file_name = Path(file_path).name

        # Получаем данные из txt_done
        txt_values = (
            row.iloc[col_map['месяц платежа']] if col_map['месяц платежа'] < len(row) else None,
            row.iloc[col_map['год платежа']] if col_map['год платежа'] < len(row) else None,
            row.iloc[col_map['общий ежемесячный платеж']] if col_map['общий ежемесячный платеж'] < len(row) else None,
            row.iloc[col_map['платеж в счет погашения основного долга']] if col_map['платеж в счет погашения основного долга'] < len(row) else None,
            row.iloc[col_map['платеж процентов']] if col_map['платеж процентов'] < len(row) else None
        )

        # Открываем файл через COM для поиска жёлтой строки
        excel = None
        wb = None
        try:
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(os.path.abspath(file_path))
            if wb.Sheets.Count < 2:
                warnings.append(f"Строка {idx+2}: в файле нет второго листа")
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
                continue
            ws = wb.Sheets(2)
            last_yellow = find_last_yellow_row_excel(ws)
            if last_yellow is None:
                warnings.append(f"Строка {idx+2}: жёлтые строки не найдены в {file_name}")
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
                continue
            row_to_paint = last_yellow + 1
            max_row = ws.UsedRange.Rows.Count
            if row_to_paint > max_row:
                try:
                    last_cell = ws.Cells.SpecialCells(11)
                    if last_cell and last_cell.Row >= row_to_paint:
                        max_row = last_cell.Row
                    else:
                        warnings.append(f"Строка {idx+2}: строка {row_to_paint} за пределами")
                        if wb:
                            wb.Close(SaveChanges=False)
                        if excel:
                            excel.Quit()
                        continue
                except:
                    warnings.append(f"Строка {idx+2}: строка {row_to_paint} за пределами")
                    if wb:
                        wb.Close(SaveChanges=False)
                    if excel:
                        excel.Quit()
                    continue

            # Проверяем совпадение данных
            xls_values = (
                ws.Cells(row_to_paint, 4).Value,
                ws.Cells(row_to_paint, 5).Value,
                ws.Cells(row_to_paint, 6).Value,
                ws.Cells(row_to_paint, 7).Value,
                ws.Cells(row_to_paint, 8).Value
            )
            txt_norm = [normalize_value(v) for v in txt_values]
            xls_norm = [normalize_value(v) for v in xls_values]
            fields = ['месяц', 'год', 'общий платеж', 'осн.долг', 'проценты']
            mismatch_fields = []
            # Сравниваем числовые значения с допуском 0.01 для чисел
            for i, (t, x) in enumerate(zip(txt_norm, xls_norm)):
                # Если оба числа, сравниваем с округлением до 2 знаков
                if isinstance(t, float) and isinstance(x, float):
                    if round(t, 2) != round(x, 2):
                        mismatch_fields.append(fields[i])
                elif t != x:
                    mismatch_fields.append(fields[i])
            if mismatch_fields:
                # Для вывода используем округлённые значения
                def fmt(v):
                    if isinstance(v, float):
                        return f"{round(v, 2):.2f}"
                    return str(v) if v is not None else ""
                detail = (f"Строка {idx+2}: данные не совпадают для {file_name} (строка XLS: {row_to_paint})\n"
                          f"  Ожидалось (из txt_done): месяц={fmt(txt_norm[0])}, год={fmt(txt_norm[1])}, "
                          f"общий платеж={fmt(txt_norm[2])}, осн.долг={fmt(txt_norm[3])}, проценты={fmt(txt_norm[4])}\n"
                          f"  Получено (из xls):       месяц={fmt(xls_norm[0])}, год={fmt(xls_norm[1])}, "
                          f"общий платеж={fmt(xls_norm[2])}, осн.долг={fmt(xls_norm[3])}, проценты={fmt(xls_norm[4])}\n"
                          f"  Не совпали поля: {', '.join(mismatch_fields)}")
                warnings.append(detail)
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
                continue

            # Всё хорошо – добавляем
            rows_to_paint.append({
                'txt_row': idx + 2,
                'file_path': file_path,
                'file_name': file_name,
                'row_to_paint': row_to_paint,
                'month': normalize_value(txt_values[0]),
                'year': normalize_value(txt_values[1]),
            })
            logger.info(f"  Строка {idx+2} → строка {row_to_paint} в {file_name}")

            if wb:
                wb.Close(SaveChanges=False)
            if excel:
                excel.Quit()
        except Exception as e:
            warnings.append(f"Строка {idx+2}: ошибка при обработке {file_name}: {e}")
            try:
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
            except:
                pass
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass

    # Выводим итоговую таблицу для подтверждения
    logger.info("")
    if warnings:
        logger.warning("Предупреждения при сборе данных:")
        for w in warnings:
            logger.warning(f"  {w}")

    if not rows_to_paint:
        logger.info("Нет строк для закрашивания.")
        return []

    logger.info("")
    logger.info(f"Найдено {len(rows_to_paint)} строк для закрашивания:")
    logger.info(" №  Файл                                     Строка   Месяц   Год")
    logger.info("--- ---------------------------------------- -------- ------- ------")
    for i, item in enumerate(rows_to_paint, 1):
        file_name = item['file_name']
        if len(file_name) > 40:
            file_name = file_name[:37] + "..."
        # Месяц и год могут быть float, выводим как целые
        month_str = str(int(item['month'])) if isinstance(item['month'], float) and item['month'].is_integer() else str(item['month'])
        year_str = str(int(item['year'])) if isinstance(item['year'], float) and item['year'].is_integer() else str(item['year'])
        logger.info(f" {i:>2}  {file_name:<40} {item['row_to_paint']:>8} {month_str:>7} {year_str:>4}")
    logger.info("")

    return rows_to_paint


def paint_rows(rows_to_paint, txt_done_path, progress_callback=None):
    """Выполняет закрашивание строк, предварительно создав резервные копии"""
    if not rows_to_paint:
        logger.info("Нет строк для закрашивания.")
        return

    unique_files = set()
    for item in rows_to_paint:
        unique_files.add(item['file_path'])
    unique_files = list(unique_files)

    logger.info(">>> НАЧАЛО ЗАКРАШИВАНИЯ")
    logger.info(f"Будет закрашено {len(rows_to_paint)} строк в {len(unique_files)} файлах.")

    backup_dir = create_backup_for_files(unique_files, txt_done_path)
    if backup_dir is None:
        logger.error("Не удалось создать резервные копии. Операция прервана.")
        return

    success_count = 0
    error_count = 0
    for i, item in enumerate(rows_to_paint):
        if progress_callback:
            progress_val = 0.5 + ((i + 1) / len(rows_to_paint)) * 0.45
            progress_callback(progress_val, f"закрашивание {i+1}/{len(rows_to_paint)}")

        logger.info(f"  Закрашивание: {item['file_name']} строка {item['row_to_paint']}")
        ok = paint_row_in_xls(item['file_path'], item['row_to_paint'])
        if ok:
            success_count += 1
        else:
            error_count += 1

    logger.info("")
    logger.info("=" * 80)
    logger.info(">> ОБРАБОТКА ЗАВЕРШЕНА")
    logger.info(f"Успешно закрашено: {success_count}, ошибок: {error_count}")
    logger.info("=" * 80)


# =========================
# ИНТЕРФЕЙС
# =========================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.log_queue = queue.Queue()
        setup_logger(self.log_queue)
        self.title(APP_NAME)
        self.geometry("1000x900")
        self.minsize(900, 750)
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.worker_thread = None
        self.create_backup_var = ctk.BooleanVar(value=True)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(5, weight=1)
        for i in range(5):
            self.grid_rowconfigure(i, weight=0)

        self.build_ui()
        self.after(100, self.poll_log_queue)

    def build_ui(self):
        # Верхняя панель
        self.top_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.top_frame.grid(row=0, column=0, padx=20, pady=(20,10), sticky="ew")
        self.top_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(self.top_frame, text=APP_NAME, font=ctk.CTkFont(size=28, weight="bold"), text_color="#1a1a1a").grid(row=0, column=0, sticky="w")
        logo_path = get_resource_path(LOGO_FILENAME)
        if os.path.exists(logo_path):
            try:
                img = Image.open(logo_path)
                new_h = 70
                w, h = img.size
                new_w = int(w * new_h / h)
                logo = ctk.CTkImage(light_image=img, dark_image=img, size=(new_w, new_h))
                ctk.CTkLabel(self.top_frame, image=logo, text="").grid(row=0, column=1, padx=(15,0), sticky="e")
            except: pass

        # Рамка ввода
        self.input_frame = ctk.CTkFrame(self, fg_color="#f8f8f8", border_width=1, border_color="#cccccc")
        self.input_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        self.input_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(self.input_frame, text="Файл txt_done.xlsx:", text_color="#1a1a1a").grid(row=0, column=0, padx=12, pady=(12,6), sticky="w")
        self.entry_file = ctk.CTkEntry(self.input_frame, fg_color="#ffffff", border_color="#cccccc", text_color="#1a1a1a")
        self.entry_file.grid(row=0, column=1, padx=12, pady=(12,6), sticky="ew")
        self.entry_file.insert(0, str(EXE_DIR / "txt_done.xlsx"))
        self.btn_file = ctk.CTkButton(self.input_frame, text="Обзор", width=100, fg_color="#2a6b3e", hover_color="#1e4d2e", text_color="#ffffff", command=self.choose_file)
        self.btn_file.grid(row=0, column=2, padx=12, pady=(12,6))

        ctk.CTkLabel(self.input_frame, text="Обработка: поиск строк-источников в XLS графиках и закраска их желтым цветом", justify="left", text_color="#1a1a1a").grid(row=1, column=0, columnspan=3, padx=12, pady=(6,12), sticky="w")

        # Рамка действий
        self.action_frame = ctk.CTkFrame(self, fg_color="#f8f8f8", border_width=1, border_color="#cccccc")
        self.action_frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        self.action_frame.grid_columnconfigure(3, weight=1)

        self.btn_run = ctk.CTkButton(self.action_frame, text="Запустить обработку", width=180, height=40, fg_color="#2a6b3e", hover_color="#1e4d2e", text_color="#ffffff", command=self.start_processing)
        self.btn_run.grid(row=0, column=0, padx=12, pady=12, sticky="w")

        self.btn_restore = ctk.CTkButton(self.action_frame, text="Восстановить из резервной копии", width=240, height=40, fg_color="#2a6b3e", hover_color="#1e4d2e", text_color="#ffffff", command=self.restore_from_backup)
        self.btn_restore.grid(row=0, column=1, padx=12, pady=12, sticky="w")

        self.btn_clear = ctk.CTkButton(self.action_frame, text="Очистить лог", width=140, fg_color="#2a6b3e", hover_color="#1e4d2e", text_color="#ffffff", command=self.clear_log)
        self.btn_clear.grid(row=0, column=2, padx=12, pady=12, sticky="w")

        self.status_label = ctk.CTkLabel(self.action_frame, text="Статус: готово", anchor="w", text_color="#1a1a1a")
        self.status_label.grid(row=0, column=3, padx=12, pady=12, sticky="ew")

        # Галочка бэкапа
        self.backup_frame = ctk.CTkFrame(self, fg_color="#f8f8f8", border_width=1, border_color="#cccccc")
        self.backup_frame.grid(row=3, column=0, padx=20, pady=10, sticky="ew")
        self.backup_checkbox = ctk.CTkCheckBox(self.backup_frame, text="Создать резервные копии перед запуском", variable=self.create_backup_var, fg_color="#2a6b3e", hover_color="#1e4d2e", checkmark_color="#ffffff", text_color="#1a1a1a")
        self.backup_checkbox.grid(row=0, column=0, padx=12, pady=8, sticky="w")

        # Прогресс
        self.progress = ctk.CTkProgressBar(self, progress_color="#2a6b3e", fg_color="#e0e0e0")
        self.progress.grid(row=4, column=0, padx=20, pady=(10,5), sticky="ew")
        self.progress.set(0)

        # Лог (моноширинный шрифт)
        self.log_box = ctk.CTkTextbox(self, wrap="word", fg_color="#ffffff", border_color="#cccccc", text_color="#1a1a1a", border_width=1, font=ctk.CTkFont(family="Consolas", size=12))
        self.log_box.grid(row=5, column=0, padx=20, pady=(0,20), sticky="nsew")
        self.log_box.configure(state="disabled")
        self.log_box.tag_config("info", foreground="#1a1a1a")
        self.log_box.tag_config("warning", foreground="#b8860b")
        self.log_box.tag_config("error", foreground="#cc0000")
        self.log_box.tag_config("critical", foreground="#cc0000")

    def choose_file(self):
        path = filedialog.askopenfilename(title="Выберите txt_done.xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if path:
            self.entry_file.delete(0, "end")
            self.entry_file.insert(0, path)

    def clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def append_log(self, text):
        self.log_box.configure(state="normal")
        if "| ERROR |" in text or "ОШИБКА" in text.upper():
            tag = "error"
        elif "| WARNING |" in text or "WARNING" in text.upper():
            tag = "warning"
        elif "| CRITICAL |" in text or "КРИТИЧЕСКАЯ" in text.upper():
            tag = "critical"
        else:
            tag = "info"
        self.log_box.insert("end", text + "\n", tag)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def update_progress(self, value, status_text=None):
        self.progress.set(value)
        if status_text:
            self.status_label.configure(text=f"Статус: {status_text}")

    def poll_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                self.append_log(msg)
                if "СТАРТ" in msg:
                    self.update_progress(0.05, "сбор данных...")
                elif "Найдено" in msg and "строк для закрашивания" in msg:
                    self.update_progress(0.5, "ожидание подтверждения...")
                elif "ОБРАБОТКА ЗАВЕРШЕНА" in msg:
                    self.update_progress(1.0, "готово!")
                elif "ошибка" in msg.lower() and "критическая" not in msg.lower():
                    self.update_progress(0, "ошибка")
        except queue.Empty:
            pass
        self.after(100, self.poll_log_queue)

    def set_controls_state(self, enabled):
        state = "normal" if enabled else "disabled"
        self.btn_run.configure(state=state)
        self.btn_restore.configure(state=state)
        self.btn_clear.configure(state=state)
        self.btn_file.configure(state=state)
        self.entry_file.configure(state=state)
        self.backup_checkbox.configure(state=state)

    def start_processing(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return
        path = self.entry_file.get().strip()
        if not path or not Path(path).exists():
            self.append_log("Ошибка: файл не выбран или не существует")
            self.status_label.configure(text="Статус: ошибка")
            return
        self.progress.set(0)
        self.status_label.configure(text="Статус: сбор данных...")
        self.set_controls_state(False)
        self.append_log("=" * 80)
        self.append_log(">> ЗАПУСК ОБРАБОТКИ")
        self.append_log("=" * 80)

        self.worker_thread = threading.Thread(target=self.collect_worker, args=(Path(path),), daemon=True)
        self.worker_thread.start()

    def collect_worker(self, txt_done_path):
        try:
            rows_to_paint = collect_rows_to_paint(txt_done_path, progress_callback=self.update_progress)
            if not rows_to_paint:
                self.after(0, lambda: self.set_controls_state(True))
                self.after(0, lambda: self.status_label.configure(text="Статус: нет данных"))
                return
            self.after(0, lambda: self.show_confirmation(rows_to_paint, txt_done_path))
        except Exception as e:
            logger.error("КРИТИЧЕСКАЯ ОШИБКА при сборе")
            logger.error(str(e))
            logger.error(traceback.format_exc())
            self.after(0, lambda: self.update_progress(0, "ошибка"))
            self.after(0, lambda: self.set_controls_state(True))

    def show_confirmation(self, rows_to_paint, txt_done_path):
        msg = f"Будет закрашено {len(rows_to_paint)} строк в {len(set(item['file_path'] for item in rows_to_paint))} файлах.\n\nПродолжить?"
        if messagebox.askyesno("Подтверждение закрашивания", msg):
            self.append_log(">>> ПОДТВЕРЖДЕНО ПОЛЬЗОВАТЕЛЕМ")
            self.status_label.configure(text="Статус: закрашивание...")
            self.worker_thread = threading.Thread(target=self.paint_worker, args=(rows_to_paint, txt_done_path), daemon=True)
            self.worker_thread.start()
        else:
            self.append_log(">>> ОТМЕНА ПОЛЬЗОВАТЕЛЕМ")
            self.append_log("Операция отменена. Изменения не внесены.")
            self.status_label.configure(text="Статус: отменено")
            self.set_controls_state(True)
            self.progress.set(0)

    def paint_worker(self, rows_to_paint, txt_done_path):
        try:
            paint_rows(rows_to_paint, txt_done_path, progress_callback=self.update_progress)
        except Exception as e:
            logger.error("КРИТИЧЕСКАЯ ОШИБКА при закрашивании")
            logger.error(str(e))
            logger.error(traceback.format_exc())
            self.after(0, lambda: self.update_progress(0, "ошибка"))
        finally:
            self.after(0, lambda: self.set_controls_state(True))

    def restore_from_backup(self):
        backups = get_all_backups()
        if not backups:
            messagebox.showwarning("Нет резервных копий", "Резервные копии не найдены.")
            return
        selected = backups[0] if len(backups) == 1 else None
        if not selected:
            names = "\n".join([f"{i+1}. {b.name}" for i, b in enumerate(backups)])
            if not messagebox.askyesno("Выбор", f"Доступные копии:\n{names}\n\nИспользовать последнюю?"):
                return
            selected = backups[0]
        txt_done_path = Path(self.entry_file.get().strip())
        if not txt_done_path.exists():
            messagebox.showerror("Ошибка", "Файл txt_done.xlsx не найден")
            return
        try:
            df, hyperlinks, file_col_idx = load_txt_done_with_hyperlinks(txt_done_path)
            graph_paths = set()
            for row_hyperlinks in hyperlinks.values():
                if file_col_idx in row_hyperlinks:
                    p = row_hyperlinks[file_col_idx]
                    if p:
                        p = p.replace('%20', ' ')
                        if p.lower().startswith("file:///"):
                            p = p[8:].replace("/", "\\")
                        graph_paths.add(p)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось прочитать файл: {e}")
            return
        if not graph_paths:
            messagebox.showerror("Ошибка", "Нет ссылок на графики")
            return
        if not messagebox.askyesno("Подтверждение", f"Восстановить {len(graph_paths)} файлов из {selected.name}?"):
            return
        self.append_log(">> ВОССТАНОВЛЕНИЕ ИЗ РЕЗЕРВНОЙ КОПИИ")
        success = restore_from_backup(selected, txt_done_path, graph_paths)
        if success:
            messagebox.showinfo("Готово", "Файлы восстановлены")
            self.append_log("Восстановление успешно")
        else:
            messagebox.showerror("Ошибка", "Ошибки при восстановлении, см. лог")
            self.append_log("Восстановление с ошибками")


if __name__ == "__main__":
    app = App()
    app.mainloop()